#===============================================================================
# Code handling the title corpus for TAL
# Author : Damien Gouteux
# Last updated : 03 April 2018
# Technologies : Python, Excel
#===============================================================================

#-------------------------------------------------------------------------------
# Import
#-------------------------------------------------------------------------------

# standard
import os
import os.path
import json
import datetime
import sys
import zipfile
from xml.sax.saxutils import escape, unescape

# external
import openpyxl
from stanfordcorenlp import StanfordCoreNLP

#-------------------------------------------------------------------------------
# Global functions
#-------------------------------------------------------------------------------

def save_to_graph(name, values):
    return
    plt.bar(range(len(values)), values.values(), align="center")
    plt.xticks(range(len(values)), list(values.keys()))
    plt.legend((name), 'upper left')
    plt.autoscale(True)
    plt.grid(True)
    #plt.show()
    plt.savefig(name + '.png')

def segment_string(string):
    """
        Used in constructor Title.
    """
    words = []
    start = 0
    end = 0
    length = 0
    for c in string:
        if start == end: # we are not in a word
            if c.isspace() or c in ["“", '"', "'", '’', '.', '«', '»', '°', '(', ')', '/', '\\', ':', '[', ']', ',', '•', '″', '…', '„', '‘']:
                start += 1
                end += 1
            else:
                end += 1
        else: # we are in a word
            if c.isspace() or c in ["“", '"', "'", '’', '.', '«', '»', '°', '(', ')', '/', '\\', ':', '[', ']', ',', '•', '″', '…', '„', '‘']:
                words.append((start, end))
                start = end + 1
                end = start
            else:
                end += 1
    if start != end:
        words.append((start, end))
    return words

#-------------------------------------------------------------------------------
# Global variables
#-------------------------------------------------------------------------------

# if a title has twice a not alphanumeric, it is counted only one
# for this character. We don't want to know how many time there is "."
# in a title, but how many titles have at least one "." in
SPECIAL_CHAR_COUNT = {}

#-------------------------------------------------------------------------------
# Classes
#-------------------------------------------------------------------------------

class ExcelFile:

    def __init__(self, name, mode='r'):
        self.name = name
        self.mode = mode
        #self.wb = xlwt.Workbook()
        self.wb = openpyxl.Workbook()
        ws = self.wb.active
        ws.title = 'Information'
        
    def save_to_sheet(self, name, values, percent=None, test_val=None):
        # ws = wb.add_sheet(name)
        ws = self.wb.create_sheet(name)
        row = 1
        for val in sorted(values, key=values.get, reverse=True): #sorted(values.keys()):
            if test_val is None or test_val(values[val]):
                ws.cell(column=1, row=row, value=val)
                #ws.write(row, 0, val)
                nb = 2
                if type(values[val]) in [tuple, list]:
                    for v in values[val]:
                        ws.cell(column=nb, row=row, value=v)
                        #ws.write(row, nb, v)
                        nb += 1
                else:
                    ws.cell(column=nb, row=row, value=values[val])
                    #ws.write(row, nb, values[val])
                    nb += 1
                if percent is not None:
                    ws.cell(column=nb, row=row, value=values[val] / percent)
                    #ws.write(row, nb, values[val] / percent)
                row += 1

    def load(self):
        if self.mode != 'r':
            raise Exception('This ExcelFile should be on read mode.')
        wb = openpyxl.load_workbook(self.name + '.xlsx', keep_vba=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        table = {}
        for row in ws.rows:
            record = []
            for cell in row:
                record.append(cell.value)
            if len(record) > 1:
                table[record[0]] = record[1:]
        return table
    
    def save(self):
        if self.mode != 'w':
            raise Exception('This ExcelFile should be on write mode.')
        self.wb.save(filename=self.name+'.xlsx')


class Repository:
    """
        Load from disk all the files of the corpus.
        There are 3046 files, with 100 records in each.
        Total: 304 600 records. Some may be discarded by checks.
        Total after checks: 298 118
    """

    def __init__(self, path, id_discard=[], char_discard=[], nlp=None):
        "Discard must be a dict or a list of id discarded."
        self.path = path
        self.filenames = os.listdir(self.path)
        self.num_found = None
        self.num_read = 0
        self.titles = []
        self.discarded = {}
        self.id_discard = id_discard
        self.char_discard = char_discard
        self.nb_id_discard = 0
        #self.nb_char_discard = 0 & char_discard are not used
        self.nlp = nlp
    
    def count_files(self):
        return len(self.filenames)

    def count_titles(self):
        return len(self.titles)
    
    def load_all(self, step=1000):
        count = 0
        total = 0
        for i in range(0, len(self.filenames)):
            self.load_one(i)
            count += 1
            if count == step:
                total += count
                print('Total loaded:', total)
                count = 0
    
    def load_one(self, num):
        filename = self.filenames[num]
        path = os.path.join(self.path, filename)
        content = json.load(open(path, encoding='utf8'))
        response = content["response"]
        # numFound check
        if self.num_found is None:
            self.num_found = response['numFound']
        elif self.num_found != response['numFound']:
            # Tolerating +1 or +2 in response
            if self.num_found != response['numFound'] - 1 and self.num_found != response['numFound'] - 2:
                print('[ERROR]')
                print(f'self.num_found = {self.num_found}')
                print(f'response["numFound"] = {response["numFound"]}')
                print(f'in doc {filename}')
                raise Exception('numFound not corresponding to the previously found.')
        # counting
        docs = response['docs']
        self.num_read += len(docs)
        for doc in docs:
            try:
                if doc['docid'] not in self.id_discard:
                    self.titles.append(Title.from_raw_json(doc, filename, self.nlp))
                else:
                    self.nb_id_discard += 1
            except KeyError as ke:
                kes = str(ke)
                if kes not in self.discarded:
                    self.discarded[kes] = []
                self.discarded[kes].append(str(doc['docid']) + ' in ' + filename)
    
    def __str__(self):
        return 'Repository ' + self.path + ' (' + str(self.num_read) + ')'

    def __repr__(self):
        return str(self)

    def dump(self, output='json', minimize=True, index='docid',makezip=False):
        if not os.path.isdir('output_dump_repo'):
             os.makedirs('output_dump_repo')
        filename = 'dump_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        if output == 'json':
            data = {}
            for t in self.titles:
                title_data = t.to_json()
                if index == 'docid':
                    data[t.docid] = title_data
                else:
                    raise Exception('Index not known: ' + index)
            if minimize:
                indent=None
            else:
                indent='    '
            full_filename = 'output_dump_repo' + os.sep + filename + '.' + output
            outfile = open(full_filename, encoding='utf-8', mode='w')
            json.dump(data, outfile, ensure_ascii=False, indent=indent)
            outfile.close()
        elif output == 'xml':
            full_filename = 'output_dump_repo' + os.sep + filename + '.' + output
            outfile = open(full_filename, encoding='utf-8', mode='w')
            outfile.write('<notices>\n')
            for t in self.titles:
                try:
                    outfile.write(t.to_xml())
                except Exception as problem:
                    print(problem)
                    print(t.title)
            outfile.write('</notices>')
            outfile.close()
        else:
            raise Exception('Not known format: ' + output)
        if makezip:
            out = zipfile.ZipFile('output_dump_repo' + os.sep + filename + '_' + output + '.zip', mode='w')
            out.write(full_filename)
            out.close()
    
    def discarded_info(self):
        for key in self.discarded:
            print(f'    {len(self.discarded[key])} because of missing key {key}')
        print(f'    {len(self.discard)} where also discarded as specified by the parameter of the constructor.')


class Domain:
    """
        A scientific domain.
        Domains are structured in a tree with multiple roots (= 0 level nodes).
    """
    
    def __init__(self, name, level):
        self.name = name
        self.level = level
        self.titles = []
        self.children = {}
        self.parent = None
        self.fullname = '{:02}'.format(self.level) + ' ' + self.name
    
    def display(self, nb, out=sys.stdout):
        ttl = len(self.titles)
        out.write('    ' * (self.level + 1) + f"{self.level}.{nb} {self}\n")
        i = 1
        for child in sorted(self.children.keys()):
            ttl += self.children[child].display(i, out)
            i += 1
        return ttl
    
    @staticmethod
    def register_list(domain_list, title):
        domains = []
        for domain_code in domain_list:
            domains.append(Domain.register_one(domain_code))
        domains[-1].titles.append(title) # we link the title and the domain only once
        return [dom for dom in domains if dom.level == 0] # we retains only the first level domain to put into title#domains
    
    ROOTS = {}
    
    @staticmethod
    def register_one(domain_code):
        element = domain_code.split('.')
        level = int(element[0])
        domain = None
        if level == 0:
            if len(element) != 2:
                raise Exception('A 0-level domain can only have one following name.')
            if element[1] not in Domain.ROOTS:
                Domain.ROOTS[element[1]] = Domain(element[1], 0)
            domain = Domain.ROOTS[element[1]]
        else: # level > 0
            parent = Domain.ROOTS[element[1]]
            for i in range(2, len(element)):
                if element[i] not in parent.children:
                    parent.children[element[i]] = Domain(element[i], level)
                parent = parent.children[element[i]]
            domain = parent
        return domain

    def __str__(self):
        return self.name.upper() + ' (lvl=' + str(self.level) + ', chld=' + str(len(self.children)) + ', ttl=' + str(len(self.titles)) + ')'


class Author:
    """
        An author of scientific work.
    """
    
    ALL_AUTHORS = {}
    
    def __init__(self, name):
        self.name = name
        self.titles = []

    @staticmethod
    def register(name, title):
        if name not in Author.ALL_AUTHORS:
            Author.ALL_AUTHORS[name] = Author(name)
        Author.ALL_AUTHORS[name].titles.append(title)
        return Author.ALL_AUTHORS[name]

    def __repr__(self):
        return self.name


class Title:
    
    def __init__(self):
        self.words = []
        self.pos_tags = []
        self.authors = []
    
    @staticmethod
    def from_raw_json(self, dic, filename, nlp=None, count_special_char=False):
        t = Title()
        # Atomic values
        t.docid = dic['docid']
        t.kind = dic['docType_s']
        t.date = dic['modifiedDateY_i']
        t.filename = filename
        # Authors
        author_list = dic['authFullName_s']
        for author in author_list:
            t.authors.append(Author.register(author, t))
        # Domains
        t.raw_domains = dic['domain_s']
        t.domains = Domain.register_list(self.raw_domains, t)
        # Title
        title_list = dic['title_s']
        t.title = title_list[0]
        #self.raw_title & self.filter() ?
        t.char_count = len(t.title)
        if nlp is not None:
            t.words = nlp.word_tokenize(t.title)
            t.pos_tags = nlp.pos_tag(t.title)
        else:
            t.words = segment_string(t.title)
            t.pos_tags = []
        t.word_count = len(self.words)
        t.special_char = False
        t.special_char_count = {}
        if count_special_char:
            for c in self.title:
                if not c.isalnum() and not c.isspace():
                    t.special_char = True
                    if c not in t.special_char_count:
                        t.special_char_count[c] = 1
                        if c not in SPECIAL_CHAR_COUNT:
                            SPECIAL_CHAR_COUNT[c] = 0
                        SPECIAL_CHAR_COUNT[c] += 1
                    else:
                        t.special_char_count[c] += 1
        # Lang
        lang_list = dic['language_s']
        if len(lang_list) > 1:
            raise Exception('Too many langs')
        t.lang = lang_list[0]
        if t.lang != 'fr':
            raise KeyError('lang')
        return t
    
    @staticmethod
    def from_xml(elem):
        t = Title()
        # Atomic values
        for child in elem:
            if child.tag == 'id':
                t.docid = int(child.text)
            elif child.tag == 'type':
                t.kind = child.text
            elif child.tag == 'date':
                t.date = child.text
            elif child.tag == 'title':
                if child.text is not None:
                    t.title = unescape(child.text)
                else:
                    t.title = ''
            elif child.tag == 'words':
                for word in child:
                    t.words.append(word.text)
            elif child.tag == 'pos_tags':
                for tag in child:
                    t.pos_tags.append(tag.text)
            elif child.tag == 'authors':
                for author in child:
                    t.authors.append(unescape(author.text))
            elif child.tag == 'domain':
                for domain in child:
                    t.domains.append(domain.text)
        return t
    
    def to_xml(self):
        authors_xml = ''
        for a in self.authors:
            authors_xml += f'            <author>{a.name}</author>\n'
        domains_xml = ''
        for d in self.raw_domains:
            domains_xml += f'            <domain>{d}</domain>\n'
        words_xml = ''
        pos_tags_xml = ''
        for p in self.pos_tags:
            words_xml += f'            <word>{escape(p[0])}</word>\n'
            pos_tags_xml += f'            <pos_tag>{p[1]}</pos_tag>\n'
        data = """    <notice>
        <id>{0}</id>
        <type>{1}</type>
        <date>{2}</date>
        <title>{3}</title>
        <words>\n{4}        </words>
        <pos_tags>\n{5}        </pos_tags>
        <authors>\n{6}        </authors>
        <domains>\n{7}        </domains>
    </notice>\n""".format(self.docid, self.kind, self.date, escape(self.title), words_xml, pos_tags_xml, authors_xml, domains_xml)
        return data
    
    def to_json(self):
        authors = []
        for a in self.authors:
            authors.append(a.name)
        #domains = []
        #for d in self.domains:
        #    domains.append(d.fullname)
        data = {
            'id' : self.docid,
            'type' : self.kind,
            'date' : self.date,
            'title' : self.title,
            'authors' : authors,
            'domains' : self.raw_domains,
            'words' : self.words,
            'pos_tags' : self.pos_tags,
            #'words' : self.words,
            #'count' : self.word_count,
            #'special' : self.special_char,
            #'lang' : self.lang
        }
        return data

##    def filter(self):
##        for c in self.raw_title:
##            if not c.isalpha() and not c.isdigit() and c not in [
##                '"', "'", '’', '‘', '«', '»', '‟', '„', '´', '`', 
##                '(', ')', '[', ']', '{', '}',
##                '*', '/', '//', '+', '-', '%', '=', '×', '^', '∞',
##                '—', '—', # the second is longer
##                '&', '®', '©', '™', '\\', '@', '#', '|', '•', '†', '·', '●',
##                '€', '$', '£', '¢',
##                ',', ';', ':', 
##                ' ', ' ', # the second is an unbreakable space
##                '.', '!', '?', '…',
##                '~']:
##                print(c, ord(c), ':', self.raw_title)
##        return self.raw_title
    
    def get_words(self):
        for delim in self.words:
            yield t.title[delim[0]:delim[1]]
    
    def __repr__(self):
        return f"{self.docid}" # in {self.filename}"


class Statistic:

    def __init__(self, repo):
        self.repo = repo

    def count_values(self, key):
        values = {}
        for t in repo.titles:
            val = getattr(t, key)
            if val not in values:
                values[val] = 1
            else:
                values[val] += 1
        return values

    def count_length(self, key, threshold=None, export=None):
        sums = {}
        for t in repo.titles:
            val = len(getattr(t, key))
            if val not in sums:
                sums[val] = 1
            else:
                sums[val] += 1
        return sums

    def count_word_n(self, index):
        values = {}
        for t in repo.titles:
            if index >= 0 and index < len(t.words):
                delim = t.words[index]
                word = t.title[delim[0]:delim[1]]
                if word not in values:
                    values[word] = 1
                else:
                    values[word] += 1
        return values

    def where_is_it(self, what):
        values = {}
        for t in repo.titles:
            find = t.title.find(what)
            if find != -1:
                if find not in values:
                    values[find] = 1
                else:
                    values[find] += 1
        return values
    
    def count_word_after(self, after):
        values = {}
        for t in repo.titles:
            find = t.title.find(after)
            if find != -1:
                for w in t.words:
                    if w[0] > find + len(after):
                        word = t.title[w[0]:w[1]]
                        if word not in values:
                            values[word] = 1
                        else:
                            values[word] += 1
                        break
        return values
    
    def count_values_n(self, key, index):
        values = {}
        for t in repo.titles:
            att = getattr(t, key)
            if index >= 0 and index < len(att):
                val = att[index]
                if val not in values:
                    values[val] = 1
                else:
                    values[val] += 1
        return values
    
    def select(self, **keyval):
        results = []
        for t in repo.titles:
            ok = True
            for key, val in keyval.items():
                if getattr(t, key) != val:
                    ok = False
                    break
            if ok:
                results.append(t)
        return results

    def info(self, idv):
        for t in repo.titles:
            if t.docid == idv:
                print('docid =', t.docid)
                print('kind =', t.kind)
                print('date =', t.date)
                print('filename =', t.filename)
                print('title =', t.title)
                print('char_count=', t.char_count)
                print('word_count=', t.word_count)
                print('lang=', t.lang)
                print('authors=')
                for auth in t.authors:
                    print('   ', auth)
                print('domains=')
                if t.domains is not None:
                    print('   ', t.domains)
                    #for dom in t.domains:
                    #    print('   ', dom)
                else:
                    print('    None')
                for i in range(0, len(self.words)):
                    if i < len(self.pos_tags):
                        print(i, self.words[i], self.pos_tags[i])
                    else:
                        print(i, self.words[i])
                break

import xml.etree.ElementTree as ET

class Corpus:

    def __init__(self, filepath):
        start = datetime.datetime.now()
        self.titles = {}
        #tree = ET.parse(filepath)
        #root = tree.getroot()
        for event, elem in ET.iterparse(filepath, events=('end',)):
            if event == 'end':
                if elem.tag == 'notice':
                    t = Title.from_xml(elem)
                    self.titles[t.docid] = t
                    elem.clear()
        delta = datetime.datetime.now() - start
        print('Nb titles:', len(self.titles))
        print(f"Loaded in {delta}.")


if __name__ == '__main__':

    start = datetime.datetime.now()

    # debug
    #f = open(r'output_dump_repo\dump_2018-04-07 00-01-04.xml', mode='r', encoding='utf-8')
    ##f = open(r'output_dump_repo\dump.xml', mode='r', encoding='utf-8')
    ##lines = f.readlines()
    ##f.close()
    ##print(lines[13347945])
    ##lines[13347945] = lines[13347945].replace('&', '&amp;')
    #lines[4443448] = lines[4443448].replace('&', '&amp;')
    ##print(lines[13347945])
    ##f = open(r'output_dump_repo\dump.xml', mode='w', encoding='utf-8')
    ##for line in lines:
    ##  f.write(line)
    ##f.close()
    
    c = Corpus(r'output_dump_repo\dump.xml')
    exit()
    
    # Starting the server
    
    NLP = StanfordCoreNLP('http://localhost', port=9000, lang='fr')
        
    # Input data

    excel = ExcelFile(name=r'io_english_titles\english_title_man', mode='r')
    english_titles = excel.load()
    del excel
    english_titles = { k : v for k, v in english_titles.items() if v[2] != 'fr'}
    
    repo = Repository('corpus-3046-files-2018-02-20-197-Mo', id_discard=english_titles, nlp=NLP)
    print('Nb files :', repo.count_files())
    repo.load_all()
    #repo.load_one(0)
    print('Nb titles:', repo.count_titles())
    stat = Statistic(repo)

    # Dump

    #repo.dump(output='json', minimize=True, index='docid')
    repo.dump(output='json', minimize=False, index='docid', makezip=True)
    repo.dump(output='xml', makezip=True)
    exit()
    # Result file

    excel = ExcelFile(name='results', mode='w')
    
    # Atomic values

    by_date = stat.count_values('date')
    excel.save_to_sheet('Date | nb', by_date, repo.count_titles())
    save_to_graph('date', by_date)

    by_kind = stat.count_values('kind')
    excel.save_to_sheet('Type | nb', by_kind, repo.count_titles())
    save_to_graph('kind', by_kind)

    by_lang = stat.count_values('lang')
    excel.save_to_sheet('Lang | nb', by_lang, repo.count_titles())
    del by_lang

    by_char_count = stat.count_values('char_count')
    excel.save_to_sheet('Char Count | nb', by_char_count, repo.count_titles())
    save_to_graph('char_count', by_char_count)
    del by_char_count

    by_sc = stat.count_values('special_char')
    excel.save_to_sheet('Special char in title | nb', by_sc, repo.count_titles())
    del by_sc
    # Not alpha numeric char
    excel.save_to_sheet('Special char | nb', SPECIAL_CHAR_COUNT)

    by_word_count = stat.count_values('word_count')
    excel.save_to_sheet('Word Count | nb', by_word_count, repo.count_titles())
    save_to_graph('word_count', by_word_count)

    first_word = stat.count_word_n(0)
    excel.save_to_sheet('First word | nb', first_word, repo.count_titles())

    where_is_double_point = stat.where_is_it(':')
    excel.save_to_sheet('Place of column | nb', where_is_double_point, repo.count_titles())

    first_word_after_double_point = stat.count_word_after(':')
    excel.save_to_sheet('First word after column | nb', first_word_after_double_point, repo.count_titles())

    # Author
    print('\n----- Authors ------\n')
    print('    There is', len(Author.ALL_AUTHORS), 'authors.\n')
    nb_authors_by_length = {}
    print('    nb articles (Y) => authors (X) : There is X authors having Y articles')
    for name, author in Author.ALL_AUTHORS.items():
        nb = len(author.titles)
        if nb not in nb_authors_by_length:
            nb_authors_by_length[nb] = 0
        nb_authors_by_length[nb] += 1
    excel.save_to_sheet('Author production | nb', nb_authors_by_length)
    #for nb in sorted(nb_authors_by_length.keys()):
    #    print('   ', nb, ':', nb_authors_by_length[nb])
    #print()
    for name, author in Author.ALL_AUTHORS.items():
        if len(author.titles) >= 190:
            print('   ', name, 'has', len(author.titles), 'publications.')
    print()
    print('    nb authors (Y) => articles (X) (there is X articles having Y authors)')
    stat.count_length('authors', threshold=6)

    # Domain
    print('\n----- Domains -----\n')
    print('    There is', len(Domain.ROOTS), 'roots domain.')
    out = open('domains.txt', mode='w', encoding='utf8')
    i = 1
    nb = 0
    for _, dom in Domain.ROOTS.items():
        nb += dom.display(i, out) #print('    ', i, '. ', dom, sep='')
        i += 1
    out.close()
    print('    There is', nb, 'publications linked to the domains.')

    DOMAIN_COUNT = {}
    for t in repo.titles:
        for d in t.domains:
            if d.name not in DOMAIN_COUNT:
                DOMAIN_COUNT[d.name] = 1
            else:
                DOMAIN_COUNT[d.name] += 1
    excel.save_to_sheet('Domain | nb', DOMAIN_COUNT, repo.count_titles())

    # Titre anglais
    SEARCH_ENGLISH_TITLES = False
    ENGLISH_TITLES = {}
    if SEARCH_ENGLISH_TITLES:
        print('\n----- English titles -----\n')
        start_english = datetime.datetime.now()
        for t in repo.titles:
            en = 0
            for w in t.get_words():
                w = w.lower()
                if w in ['on', 'and', 'a', 'in', 'the', 'und']:
                    en += 1
            try:
                lg = langdetect.detect(t.title)
            except langdetect.lang_detect_exception.LangDetectException:
                lg = 'err'
            if en > 0:
                ENGLISH_TITLES[t.docid] = (en, lg, t.title)
        english_wb = xlwt.Workbook()
        save_to_excel(english_wb, 'Id | title', ENGLISH_TITLES)
        english_wb.save('english_title.xls')
        end_english = datetime.datetime.now()
        delta = end_english - start_english
        print(f"    Search of english titles took {delta}.")
    
    # Generation of the Lexicon
    GENERATE_LEXICON = True
    WORDS = {}
    TOTAL_WORD = 0
    if GENERATE_LEXICON:
        print('\n----- Generate Lexicon -----\n')
        start_lexicon = datetime.datetime.now()
        for t in repo.titles:
            for w in t.get_words():
                w = w.lower() # preformating
                if w not in WORDS:
                    WORDS[w] = 0
                WORDS[w] += 1
                TOTAL_WORD += 1
        # lexicon_wb = xlwt.Workbook()
        lexicon_wb = ExcelFile(name=r'output_lexicon\lexicon', mode='w')
        # 65536 is the maxium in a xls files. So I have to skip some
        def test_freq(freq):
            return freq > 10
        #save_to_excel(lexicon_wb, 'Word | nb', WORDS, TOTAL_WORD, test_val=test_freq)
        #lexicon_wb.save('lexicon.xls')
        lexicon_wb.save_to_sheet('Word | nb', WORDS, TOTAL_WORD, test_val=test_freq)
        lexicon_wb.save()
        end_lexicon = datetime.datetime.now()
        delta = end_lexicon - start_lexicon
        print(f"    Lexicon generation took {delta}.")

    # Generation of titles with POS tagging
    
    # Request
    print('\n----- Request -----\n')
    res = stat.select(char_count=0)
    print('    Length of selection of char_count == 0:', len(res))
    for r in res:
        print('      ', r)
    res = stat.select(char_count=1)
    print('    Length of selection of char_count == 1:', len(res))
    for r in res:
        print('      ', r)
    print('\n----- End -----\n')

    excel.save()

    end = datetime.datetime.now()

    delta = end - start

    print(f"    Script has ended [{delta} elapsed].")

    print('\n----- Discarded info -----\n')
    
    repo.discarded_info()

