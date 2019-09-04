#
# This file handles all the building of the corpus.
# 339687 titres
#
# It requires :
#   resources\recodage-josette.tsv
#   resources\LTES-Tutin-2007.txt
#   data\total-articles-HAL.tsv => Title metadata
#   data\output_tal_01.txt      => TAL data
#   data\output_tal_02.txt      => TAL data
#   data\output_tal_03.txt      => TAL data
#   data\output_tal_04.txt      => TAL data
#   data\output_tal_05.txt      => TAL data
#   data\output_tal_06.txt      => TAL data
#
# It produces :
#   corpus of titles following the datamodel (in memory)
#   unknown_lemma.txt           => list of all the unknown lemma
#
# Summary :
#   1. Imports
#   2. Handling of simplification of Domains
#   3. Data model for Words and Titles
#   4. Read titles metadata
#   5. Reading Talismane data file and updating titles
#   6. Utils
#       find_title(attr, value, stop_on_first=True, listing=True)
#       pprint(dic)
#   7. Model for results
#   8. Stats
#       select  Find ALL titles matching the key values, return a dict
#       find    Find N titles matching the key values, return a list (based on find)
#       count   Count the number of titles matching the key values
#       stat    Make stats on one or more keys
#       Exemple : 
#            res = stat(['roots.0.pos', 'domain'])
#            res_agg = aggregate(res)               Aggregate NC+NPP=>NOUN / Vx=>VERB / P+P+D=>PREP DEPRECATED : USE roots.0.pos:agg for a direct result
#            by_dom(res_agg)
#   9. Boot

#-------------------------------------------------
# Imports
#-------------------------------------------------

# Standard
import datetime
from enum import Enum
from importlib import reload # Dynamic code
import itertools # cribble
import copy
import math # for sqrt

# Project
import whiteboard as wb # Dynamic code

# Other
import openpyxl # only pour make_lexique
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

#-------------------------------------------------
# Handling lexiques
#-------------------------------------------------

TUTIN = []
ftut = open("resources\\LTES-Tutin-2007.txt", mode='r', encoding='utf8')
fdat = ftut.readlines()
for line in fdat:
    if line.startswith('--'): continue
    dat = line.split(';')
    name = dat[0].rstrip()
    TUTIN.append(name)

#-------------------------------------------------
# Handling of simplification of Domains
#-------------------------------------------------

# Loading recoding dictionary
recode=dict()

def load_recoding_table(debug=False):
    global recode
    entree = open("resources\\recodage-josette.tsv", mode="r", encoding="UTF8")
    header = entree.readline()
    for line in entree:
        line=line.rstrip("\n")
        t=line.split("\t")
        recode[t[0]]=t[3]
    entree.close()

def recode_domain(title):
    champs = title.domains.split(",")
    dom=""
    for c in champs:
        if (c.startswith("0") or c.startswith("1")) and c != "0.shs":
            if c in recode:
                return recode[c] #first domain only
            else:
                raise Exception('Recoding domain failed for title: ' + title.idt)
    return 'NONE'

#-------------------------------------------------
# Data model for Words and Titles
#-------------------------------------------------

class Word:

    unknown_lemma = {}
    
    def __init__(self, idw, form, lemma, pos, info, gov, dep):
        self.form = form
        if lemma == '_':
            self.lemma = f"?{form}"
            Word.unknown_lemma[form + ':::' + pos] = ''
        else:
            self.lemma = lemma
        self.idw = idw
        self.pos = pos
        self.info = info
        self.gov = gov # = index in title.words + 1 (start at 1, 0 = no gov) /!\ if restarts !
        self.dep = dep

    def __str__(self):
        return f"{self.form}"
    
    def __repr__(self):
        return f"<Word {self.form, self.lemma, self.pos}>"

    @classmethod
    def write_unknown_lemma(cls):
        out = open('unknown_lemma.txt', mode='w', encoding='utf8')
        for wp in cls.unknown_lemma:
            out.write(wp + '==>?\n')
        out.close()


class Title:

    def __init__(self, idt, year, typ, domains, authors, text):
        self.idt = idt
        self.year = int(year)
        self.typ = typ
        self.domains = domains
        self.domain = recode_domain(self)
        self.authors = authors
        self.nb = authors.count(',') + 1
        self.text = text
        self.words = []
        self.restarts = []
        self.nb_restarts = 0
        self.nb_parts = 0 # nb_parts = nb_restarts + 1
        # updated in stats, must be called afer words is set
        self.nb_roots = 0
        self.roots = []
        self.len_with_ponct = 0
        self.len_without_ponct = 0
        self.nb_segments = 0
        self.segments = []
        self.roots_by_segments = 'NOVAL'
        # combi
        self.parts_segments = 'NOVAL'
    
    def __repr__(self):
        return f"<Title |{self.text[:20]}| #{self.typ} @{self.year} D{self.domain} Len({self.len_without_ponct} +{self.len_with_ponct - self.len_without_ponct}) Seg({self.nb_segments}) Root({self.nb_roots})>"

    def __str__(self):
        return f"{self.text}"

    def info(self):
        for i, w in enumerate(self.words):
            print(f"{i:2d}. {w.idw:3d} {w.form:16} {w.pos:5} {w.lemma:16} {w.gov:3d}, {w.dep:5}")
    
    ponct_no_seg = {}
    
    def init(self, words, restarts):
        self.words = words
        self.restarts = restarts
        self.nb_restarts = len(restarts)
        self.nb_parts = self.nb_restarts + 1
        current_seg_nb_roots = 0
        self.roots_by_segments = []
        for cpt, w in enumerate(self.words):
            if is_root(w):
                self.nb_roots += 1
                self.roots.append(cpt)
                current_seg_nb_roots += 1
            if w.pos != 'PONCT':
                self.len_without_ponct += 1
            elif is_seg(w):
                self.nb_segments += 1
                self.segments.append(cpt)
                self.roots_by_segments.append(current_seg_nb_roots)
                current_seg_nb_roots = 0
            else:
                if w.form in Title.ponct_no_seg:
                    Title.ponct_no_seg[w.form] += 1
                else:
                    Title.ponct_no_seg[w.form] = 1
            self.len_with_ponct += 1
        # Adding one if the last word is not seg
        if not is_seg(self.words[-1]):
            self.nb_segments += 1
            self.roots_by_segments.append(current_seg_nb_roots)
        # Check 1
        if sum(self.roots_by_segments) != len(self.roots):
            raise Exception("A root has no segment!")
        # Check 2
        if len(self.roots_by_segments) != self.nb_segments:
            raise Exception('Invalid count roots_by_segment', self.idt)
        self.parts_segments = f"{self.nb_parts:1d}:{self.nb_segments:1d}"
        self.roots_by_segments = ':'.join([str(x) for x in self.roots_by_segments])

#-------------------------------------------------
# Read titles metadata
#-------------------------------------------------

def read_titles_metadata(file_name):
    """Create the titles from a TSV file"""
    print('[RUN ] --- read_titles_metadata @' + file_name)
    start_time = datetime.datetime.now()
    print('[INFO] --- Started at', start_time)
    titles = {}
    try:
        file = open(file_name, encoding='utf8', mode='r')
        content = file.readlines()
    except FileNotFoundError:
        print('File not found.')
    except UnicodeDecodeError:
        print('Invalid utf-8 format.')
    else:
        print('[INFO] ---', len(content), 'lines.') # 339687
        for line in content:
            elements = line.split('\t')
            idt = elements[0]
            year = elements[1]
            typ = elements[2][1:-1]
            domains = elements[3][1:-1]
            authors = elements[4][1:-1]
            text = elements[5][1:-1]
            t = Title(idt, year, typ, domains, authors, text)
            titles[t.idt] = t
        file.close()
    print('[INFO] ---', len(titles), 'titles.') # 339687
    end_time = datetime.datetime.now()
    print('[INFO] --- Ending at', end_time)
    delta = end_time - start_time
    print(f"[INFO] --- Script has ended [{delta} elapsed].\n")
    return titles

#-------------------------------------------------
# Reading Talismane data file and updating titles
#-------------------------------------------------

def read_update_from_talismane_data(titles, file_name):
    global titles_with_more_than_one_paragraph
    """Update the titles with Talismane informations"""
    print('[RUN ] --- read_update_from_talismane_data @ ' + file_name)
    start_time = datetime.datetime.now()
    print('[INFO] --- Started at', start_time)
    # read
    try:
        file = open(file_name, encoding="utf8", mode="r")
    except FileNotFoundError:
        raise Exception("File not found: " + file_name)
    else:
        lines = file.readlines()
        print('[INFO] --- Lines read:', len(lines))
    # parcours
    updated = 0
    for index in range(len(lines)):
        line = lines[index]
        if line.startswith('<title id="'):
            idt = line[11:len(line)-3]
            words = []
            prev_idw = None
            restarts = []
            index += 1
            line = lines[index]
            while not line.startswith('<title id="'):
                elements = line.split('\t')
                if len(elements) == 10:
                    idw = int(elements[0])
                    if prev_idw is not None and idw < prev_idw:
                        restarts.append(len(words))
                    prev_idw = idw
                    form = elements[1]
                    lemma = elements[2]
                    typ1 = elements[3] # 4 is the same
                    info = elements[5]
                    gov = int(elements[6])
                    dep = elements[7] # 8 and 9 are the same
                    words.append(Word(idw, form, lemma, typ1, info, gov, dep))
                index += 1
                try:
                    line = lines[index]
                except IndexError:
                    break
            titles[idt].init(words, restarts)
            updated += 1
            index -= 1
        index += 1
    end_time = datetime.datetime.now()
    print(f'[INFO] --- Titles updated: {updated}')
    print(f'[INFO] --- Ending at {end_time}')
    print(f"[INFO] --- Script has ended [{end_time - start_time} elapsed].\n")
    return titles

#-------------------------------------------------
# Filtering
#-------------------------------------------------

MAX_NB_PART = 1
MIN_NB_ROOT = 1
MAX_NB_ROOT = 1
MAX_NB_SEG  = 2
ROOT_POS_OK = ['NC', 'NPP']

filter_text = f"""           We keep :
             - only the title with {MAX_NB_PART} part (restart = nb_part - 1)
             - only the title with {MIN_NB_ROOT} <= root <= {MAX_NB_ROOT}
             - the root must be of type {ROOT_POS_OK}
             - nb_segments <= {MAX_NB_SEG}
"""

def filter_titles(debug=False):
    global titles
    too_many_restart = 0
    too_many_root = 0
    too_many_seg = 0
    root_not_nc_npp = 0
    ponct_not_known = 0
    keys_to_delete = []
    for kt, t in titles.items():
        if t.restart > (MAX_NB_PART - 1):
            keys_to_delete.append(kt)
            too_many_restart += 1
        elif not MIN_NB_ROOT <= t.nb_roots <= MAX_NB_ROOT:
            keys_to_delete.append(kt)
            too_many_root += 1
        elif t.nb_segments > MAX_NB_SEG:
            keys_to_delete.append(kt)
            too_many_seg += 1
        else:
            for w in t.words:
                if is_root(w) and w.pos not in ROOT_POS_OK:
                    keys_to_delete.append(kt)
                    root_not_nc_npp += 1
                    break
                elif w.pos == 'PONCT' and not is_seg(w) and not ponct_ok(w):
                    keys_to_delete.append(kt)
                    ponct_not_known += 1
                    break
    old_len = len(titles)
    for k in keys_to_delete:
        del titles[k]
    if debug:
        print(f"[INFO] --- Starting length =     {old_len:10d}")
        print(f"[INFO] --- Too many restart =    {too_many_restart:10d}")
        print(f"[INFO] --- Too many root =       {too_many_root:10d}")
        print(f"[INFO] --- Too many seg =        {too_many_seg:10d}")
        print(f"[INFO] --- Root not nc or npp =  {root_not_nc_npp:10d}")
        print(f"[INFO] --- Ponct not known =     {ponct_not_known:10d}")
        print(f"[INFO] --- Length after filter = {len(titles):10d}")

#-------------------------------------------------
# Utils
#-------------------------------------------------

def is_root(word : Word):
    return word.gov == 0 and word.dep in ['_', 'root']

def is_seg(word : Word): # : ; . ? ! + variantes du .?!
     return word.pos == 'PONCT' and word.form in [
         ':', '.', '?', '!', '...', '…', ';', '..', '....', '?.', '?!',
         '...?', '?...', '.....', '!...', '!?', '!.', '!!!', '!!',
         '......', '??', '?..', '.?', '?!...']

def ponct_ok(word : Word):
    if word.pos != 'PONCT':
        raise Exception('Not a ponct: ' + str(word))
    elif word.form in [',', '-', '"', '(', ')', "'", '[', ']', '&', '\\',
                       '*', '‑', '}', '{', '§', '†', '|', '·', '^', '‚', '‹',
                       '›', '¡', '‛', '・', '•', '/']:
        return True
    else:
        return False


def is_top_100_signoun(word : str):
    return word in ['exemple', 'cas', 'modèle', 'résultat', 'façon', 'problème',
                      'équation', 'théorie', 'terme', 'section', 'idée', 'point',
                      'figure', 'chose', 'système', 'table', 'question', 'raison',
                      'effet', 'méthode', 'procédé', 'facteur', 'type', 'fait',
                      'principe', 'réaction', 'problème', 'approche', 'expérience',
                      'procédure', 'forme', 'condition', 'taux', 'droit', 'type',
                      'solution', 'fonction', 'changement', 'valeur', 'étude',
                      'argument', 'possibilité', 'abilité', 'différence', 'loi',
                      'série', 'zone', 'concept', 'analyse', 'conclusion', 'situation',
                      'article', 'politique', 'vue', 'réponse', 'relation', 'stratégie',
                      'conséquence', 'supposition', 'étape', 'période', 'scène', 'but',
                      'discussion', 'échec', 'essai', 'propriété', 'chapitre', 'trait',
                      'caractéristique', 'expression', 'potentiel', 'technique', 'sujet',
                      'paramètre', 'mécanisme', 'instance', 'indice', 'partie',
                      'introduction', 'test', 'rôle', 'objectif', 'coefficient',
                      'décision', 'comportement', 'intention', 'prévision',
                      'hypothèse', 'nombre', 'implication', 'avantage', 'définition',
                      'observation', 'notion', 'phénomène', 'objectif', 'mot',
                      'difficulté', 'sujet']


def p(a):
    return round(a/339687*100, 2)


# find sub roots for a 2 seg titles
def find_subroot(data):
    find_direct = 0
    find_indirect = 0
    for k, t in data.items():
        if t.roots_by_segments not in ['1:0', '0:1']:
            continue
        if t.roots_by_segments == '0:1':
            start = 0
            end = t.segments[0]
        elif t.roots_by_segments == '1:0':
            start = t.segments[0] + 1
            end = len(t.words)
        found = []
        # find if there is a secondary root and how many
        for i in range(start, end):
            w = t.words[i]
            if w.gov == t.words[t.roots[0]].idw:
                found.append(i)
        if len(found) == 0:
            w = t.words[start]
            if t.roots_by_segments == '0:1' and w.gov >= end:
                found.append(start)
                find_indirect += 1
            elif t.roots_by_segments == '1:0' and w.gov <= start:
                found.append(start)
                find_indirect += 1
        elif len(found) > 0:
            find_direct += 1
        if len(found) > 0:
            t.roots.append(found[0])
            t.roots.sort() # in order to have the root of the first seg, root of second seg
            t.roots_by_segments = '1:1'
        #    raise Exception('Too many sub racines for t=' + t.idt)
        #t.subroots = found
    return find_direct, find_indirect


#-------------------------------------------------
# Model for results
#-------------------------------------------------

class Column:

    def __init__(self, data, name, kind='num', wide=10):
        self.data = data
        self.name = name # can be a tuple (or any kind) !
        self.kind = kind
        self.total = 0
        self.max_length = 10
        self.wide = 10
    
    def update(self):
        if self.kind == 'num':
            self.total = 0
            for rec in self.data:
                self.total += self.data[rec][self.name]
            self.max_length = max(self.wide, len(self.name), len(str(self.total)))
        elif self.kind in ['str', 'tuple']:
            self.max_length = 0
            for rec in self.data:
                if len(str(self.data[rec][self.name])) > self.max_length:
                    self.max_length = len(str(self.data[rec][self.name]))
            self.max_length = max(self.max_length, len(self.name))
        elif self.kind == 'bool':
            self.max_length = 5
        else:
            raise Exception('Type of column not known, must be num, str or bool for col ' + self.name)
    
    def __str__(self):
        return f"{str(self.name):{self.max_length}} "
    
    def __getitem__(self, rec):
        return self.data[rec][self.name]


class Data:

    def __init__(self, col_defs : dict, autocount=True):
        self.columns = {}
        if 'count' not in col_defs and autocount:
            col_defs.update({'count' : 'num'})
        for n, k in col_defs.items():
            self.columns[n] = Column(self, name=n, kind=k)
        self.content = {}
        self.default_counted_columns = self.columns[next(iter(self.columns))]

    @staticmethod
    def from_dict(dic):
        "convert a dict { 'abc' : 5, 'def' : 10 } to { 'abc' : { 'count' : 5 }, 'deg' : { 'count' : 10 } }"
        d = Data({'key' : 'str', 'count' : 'num'})
        for k, v in dic.items():
            d.set_cell(k, 'key', k)
            d.set_cell(k, 'count', v)
        return d
    
    def __getitem__(self, rec):
        return self.content[rec]

    def __iter__(self):
        self.itr = iter(self.content)
        return self

    def __next__(self):
        return next(self.itr)
    
    def set_cell(self, rec, col, val):
        if rec not in self.content:
            self.content[rec] = {}
        self.content[rec][col] = val

    def get_cell(self, rec, col):
        return self.content[rec][col]
    
    def __length__(self):
        return len(self.content)
    
    def nb_columns(self):
        return len(self.col_names)

    def update(self):
        for _, col in self.columns.items():
            col.update()
    
    def count(self, rec, col = None):
        if col is None:
            col = self.default_counted_columns
        if rec in self.content:
            self.content[rec]['count'] += 1
        else:
            self.set_cell(rec, col.name, rec)
            self.set_cell(rec, 'count', 1)
    
    def display(self, sort_key = 'count', with_line = True,
                max_line = None, until_total_percent = None, min_percent = None):
        self.update()
        total_percent = 0.0
        percent = None
        len_line = 8 # for percent col
        for _, col in self.columns.items():
            len_line += len(str(col))
        # Header
        print('-' * len_line)
        for _, col in self.columns.items():
            print(str(col), end='')
            if col.name == sort_key:
                print("percent ", end='')
        print()
        print('-' * len_line)
        # Body
        lines = 0
        for k1 in sorted(self.content, key=lambda x: self.content[x][sort_key], reverse=True):
            for _, col in self.columns.items():
                print(f"{str(col[k1]):{col.max_length}} ", end='')
                if col.name == sort_key:
                    percent = (col[k1]/col.total) * 100
                    total_percent += percent
                    print(f"{percent:8.4f} ", end='')
            print()
            if with_line: print('-' * len_line)
            lines += 1
            # breakers
            if max_line is not None and lines >= max_line:
                print("... (percent indicates what's taken)")
                break
            if until_total_percent is not None and total_percent >= until_total_percent:
                print("... (percent indicates what's taken)")
                break
            if min_percent is not None and percent is not None and percent < min_percent:
                print("... (percent indicates what's taken)")
                break
        # Footer
        if not with_line: print('-' * len_line)
        for _, col in self.columns.items():
            print(str(col), end='')
            if col.name == sort_key:
                print("percent ", end='')
        print()
        for _, col in self.columns.items():
            if col.name != sort_key:
                print(len(str(col)) * ' ', end='')
            else:
                print(f"{str(col.total):{col.max_length}}", end='')
                print(f"{total_percent:8.4f}", end='')
        print()
        print('-' * len_line)
        print()

#-------------------------------------------------
# Stats
#-------------------------------------------------

stats = {}


def match_title(t, tested_keys_values):
    """
        match_title is used by find & count
        -> find retrieves titles corresponding to a set of conditions
        -> count counts the number of titles corresponding to a set of conditions
        -> match tells only if a title corresponds to a set of contions
        Keys can be :
        -> an attribute of title : nb_roots
        -> an sub-attribute of an element in a list attribute : roots.0.pos
        -> a length of an attribute : #roots
        Values can be :
        - only one (the value of the key must be corresponding to it)
        - a min and a max : through a tuple : min:max
        - a list of values : through a list : v1|v2|v3
        Exemples :
            count({'nb_roots' : 0})      only 0
            count({'nb_roots' : (0, 2)}) from 0 to 2
            count({'nb_roots' : [0, 2]}) only 0 or 2
            count({'nb_roots' : 0, 'nb_segments' : 1})
    """
    ok = True # for inclusive conditions, set this to False and set it True when a condition is reached
    for tested_key, tested_value in tested_keys_values.items():
        # Get the value
        if tested_key[0] == '#':
            attr = getattr(t, tested_key[1:])
            val = len(attr)
        elif '.' in tested_key:
            attr_key, attr_index, obj_key = tested_key.split('.') # roots.0.pos
            attr = getattr(t, attr_key)
            if int(attr_index) >= len(attr):
                raise Exception('Index not in range')
            val = getattr(t.words[attr[int(attr_index)]], obj_key) # works only for words
        else:
            val = getattr(t, tested_key)
        # Test
        if type(tested_value) in [int, str]:
            if val != tested_value:
                ok = False
        elif type(tested_value) == tuple:
            min_val = tested_value[0]
            max_val = tested_value[1]
            if not min_val <= val <= max_val:
                ok = False
        elif type(tested_value) == list:
            if val not in tested_value:
                ok = False
        else:
            raise Exception('Tested value unknown type:' + str(type(tested_value)))
    return ok


# t121a = select({'roots_by_segments' : '1:0'})
def select(tested_keys_values):
    res_list = find(tested_keys_values, len(titles), display=False)
    res_dict = {}
    for t in res_list:
        res_dict[t.idt] = t
    return res_dict


# find(nb=30, display=False, filename='t111.txt')
def find(tested_keys_values=None, nb=5, display=True, filename=None):
    if filename is not None:
        f = open(filename, mode='w', encoding='utf8')
    selected = []
    for kt, t in titles.items():
        if tested_keys_values is None or match_title(t, tested_keys_values):
            selected.append(t)
            if display:
                print(t.idt, t)
                print('-' * len(str(t)))
                t.info()
            if filename is not None:
                f.write(t.idt + ' ' + str(t) + '\n')
                f.write('-' * len(str(t)) + '\n')
                for i, w in enumerate(t.words):
                    f.write(f"{i:2d}. {w.idw:3d} {w.form:16} {w.pos:5} {w.lemma:16} {w.gov:3d}, {w.dep:5}\n")
                f.write('\n')
            if len(selected) == nb:
                break
    if filename is not None:
        f.close()
    return selected


def cribble(threshold, *keys):
    # get all the values
    vals = {}
    for k in keys:
        vals[k] = []
    for kt, t in titles.items():
        for k in keys:
            val = getattr(t, k)
            if val not in vals[k]:
                vals[k].append(val)
    # get the count for all combinaisons
    counted = {}
    total = 0
    length = len(titles)
    for combi in itertools.product(*vals.values()):
        tested = {}
        for i, k in enumerate(keys):
            tested[k] = combi[i]
        res = count(tested)
        if (res / length)* 100 >= threshold:
            counted[combi] = res
            total += res
    # display results
    print('number of combi   :', len(counted))
    print('number of titles  :', total)
    print('percent of titles :', f"{(total/length)*100:5.2}")
    # display the combi
    for combi in counted:
        print(f"{str(combi):10}", f"{counted[combi]:10d}", f"{(counted[combi]/length)*100:5.2f}")


def count(tested_keys_values=None):
    """
        Used to count the title corresponding to the values of the tested keys.
    """
    if tested_keys_values is None:
        return len(titles)
    count = 0
    for kt, t in titles.items():
        if match_title(t, tested_keys_values):
            count += 1
    return count


def uniq(key):
    """
        Used to get all the unique values of a key
    """
    vals = []
    for kt, t in titles.items():
        if getattr(t, key) not in vals:
            vals.append(getattr(t, key))
    return vals

    
def maxx(key):
    """
        Used to get the max of a numeric value of a key
    """
    val_max = 0
    for kt, t in titles.items():
        if val_max < getattr(t, key):
            val_max = getattr(t, key)
    return val_max


def minn(key):
    """
        Used to get the min of a numeric value of a key
    """
    val_min = None
    for kt, t in titles.items():
        if val_min is None or val_min > getattr(t, key):
            val_min = getattr(t, key)
    return val_min


def avg(key):
    """
        Used to get the average of a numeric value of a key
    """
    val_sum = 0
    for kt, t in titles.items():
        val_sum += getattr(t, key)
    return val_sum / len(titles)


# for calculating over a period the value (ex : x(res, 2005, 2019))
# res = stat('year')
def cumul_over_period(res, deb, end):
	 i = end
	 tot = 0
	 while i >= deb:
		 tot += res[(i,)]
		 i -= 1
	 return tot


def agg(pos):
    if pos in ['V', 'VIMP', 'VINF', 'VPP', 'VPR', 'VS']:
        return 'VERB'
    elif pos in ['NC', 'NPP']:
        return 'NOUN'
    elif pos in ['P', 'P+D']:
        return 'PREP'
    else:
        return pos


def stat(keys=None, display=True, until_total_percent=None, until_min_freq=None):
    """
        Used to make stat on one or more keys of the titles.
        - If no keys is passed, it counts the number of titles.
        - If one key is passed, it counts the number of titles for each different values of this key.
        - If more than one key is passed, it counts the number of titles for each different combinations of the values of these keys.
        - The key can be of a special format : roots.0.pos <=> t.roots[0].pos
    """
    if keys is None:
        return len(titles)
    if type(keys) == str:
        keys = [keys]
    values = {}
    for key, t in titles.items():
        vals = []
        valx = [] # only for $keys
        for k in keys:
            if '#' in k:
                attr = getattr(t, k[1:])
                vals.append(len(attr))
            elif k[0] == '$': # special keys for working with both roots as different objects. ex: $roots.lemma
                if k.startswith('$roots'):
                    mod = k.split('.')[1]
                    for r in t.roots:
                        valx.append(getattr(t.words[r], mod))
            elif '.' in k:
                attr_key, attr_index, obj_key = k.split('.') # roots.0.pos or roots.0.pos:agg
                attr = getattr(t, attr_key)
                if int(attr_index) >= len(attr):
                    vals.append('NO VAL')
                else:
                    if ':' in obj_key:
                        obj_key, mod = obj_key.split(':')
                        val = getattr(t.words[attr[int(attr_index)]], obj_key)
                        if mod == 'agg':
                            vals.append(agg(val))
                    else:
                        if attr == t.words:
                            vals.append(getattr(t.words[int(attr_index)], obj_key)) # works for words.0.lemma
                        else:
                            vals.append(getattr(t.words[attr[int(attr_index)]], obj_key)) # works only for roots.0.lemma
            else:
                vals.append(getattr(t, k))
        if len(valx) == 0:
            val = tuple(vals)
            if val not in values:
                values[val] = 0
            values[val] += 1
        else: # saving one record for each values of valx
            for v in valx:
                val = tuple([v] + vals)
                if val not in values:
                    values[val] = 0
                values[val] += 1
    total = 0
    for val in values:
        total += values[val]
    if display:
        s = f'*** {str(keys)} *** ({len(titles)})'
        print(s, '\n', '-' * len(s), sep='')
        i = 0
        cumul_percent = 0.0
        max_len = 10
        for key in values:
            if len(str(key)) > max_len: max_len = len(str(key))
        for key in sorted(values, key=values.get, reverse=True):
            i += 1
            percent = (values[key]/total)*100
            cumul_percent += percent
            d = '-'.join(map(str, key))
            print(f"{i:3d}. {d:>{max_len}} {values[key]:10d} {percent:8.4f} % {cumul_percent:6.2f} %")
            if until_total_percent is not None and cumul_percent > until_total_percent:
                print('...')
                break
            if until_min_freq is not None and values[key] <= until_min_freq:
                print('...')
                break
    return values


# Calculation of % by domain
# works for every couple of (key, domain)
# Ex : res = stat(['roots.0.pos:agg', 'domain']) ; by_dom(res)
def by_dom(data, max_pos=5):
    # Get total of key 2 (domain)
    domain_count = {}
    for k, v in data.items():
        dom = k[1]
        if dom not in domain_count:
            domain_count[dom] = v
        else:
            domain_count[dom] += v
    # Get the first 5 POS for each domain
    domain_stat = {}
    for k, v in data.items():
        key = k[0]
        dom = k[1]
        if dom not in domain_stat:
            domain_stat[dom] = {}
        domain_stat[dom][key] = round(v / domain_count[dom] * 100, 2)
    # Display
    for kdom in sorted(domain_count, key=domain_count.get, reverse=True):
        dom = domain_stat[kdom]
        print(f"{kdom:14}", end=' ')
        cpt = 0
        for key in sorted(dom, key=dom.get, reverse=True):
            v = dom[key]
            print(f"{key:5} {v:5.2f}", end='  ')
            cpt += 1
            if cpt >= max_pos: break
        print(f"{domain_count[kdom]:6d}")


# Data structure for percent & dump2sheet
# lemma::pos : lemma, pos, nb, percent
#

class Domain:

    #total_heads = 0
    total_nouns = 0
    
    def __init__(self, name):
        # il enregistre TOUS les noms
        # même dans heads, il y aura juste 0
        # donc total_heads = somme des têtes (y'en a à 0, y'en à plus de 1)
        # donc attention len(self.heads) == len(self.nouns)
        self.name = name
        self.heads = {}
        self.nouns = {}
        self.nb_heads = 0 # nb de têtes différents (sans compter le nombre d'occurrences)
        self.filter1 = 0
        self.filter2 = 0

    
    def record(self, key, head):
        if key not in self.nouns:
            self.nouns[key] = 1
            self.heads[key] = 1 if head else 0
        else:
            self.nouns[key] += 1
            if head:
                self.heads[key] += 1


    def count(self):
        self.total_heads = 0
        self.total_nouns = 0
        for h, v in self.heads.items():
            self.total_heads += v
            if v > 0: self.nb_heads += 1
        for n, v in self.nouns.items():
            self.total_nouns += v
        #Domain.total_heads += self.total_heads # ON COMPTE EN DOUBLE DES TÊTES !
        Domain.total_nouns += self.total_nouns
        #self.best_heads = sorted(self.heads, key=self.heads.get, reverse=True)
        #self.best_nouns = sorted(self.nouns, key=self.nouns.get, reverse=True)


# 1 on lex
# 2 on calc (auto appelé après lex)
# 3 on peut faire to_sheet ou select

# Stat

def fmax(serie):
    mx = 0
    for s in serie:
        if s > mx: mx = s
    return mx


def fmoy(serie):
    return sum(serie)/len(serie)


def fmed(serie):
    # Calcul de la médiane
    mid_elem = len(serie) // 2
    return sorted(serie, reverse=True)[mid_elem]


def fect(serie):
    moy = fmoy(serie)
    ect = 0
    for v in serie:
        ect += (v - moy) ** 2 # pour pas que ça fasse 0
    ect = ect / len(serie) # moyenne des écarts
    ect = math.sqrt(ect)
    return ect


def frsd(serie): # écart-type relatif en %
    return fect(serie) / fmoy(serie)


SWEET = {
    '0.chim'       : "Chimie",
    '0.info'       : "Informatique",
    '0.math'       : "Mathématiques",
    '0.phys'       : "Physique",
    '0.qfin'       : "Économie et finance quantitative",
    '0.scco'       : "Sciences cognitives",
    '0.sde'        : "Sciences de l'environnement",
    '0.sdu'        : "Planète et Univers",
    '0.sdv'        : "Sciences du Vivant",
    '1.shs.anthro' : "Anthropologie",
    '1.shs.archeo' : "Archéologie et Préhistoire",
    '1.shs.archi'  : "Architecture",
    '1.shs.art'    : "Art et histoire de l'art",
    '1.shs.autre'  : "Autres",
    '1.shs.droit'  : "Droit",
    '1.shs.edu'    : "Éducation",
    '1.shs.geo'    : "Géographie",
    '1.shs.gestion': "Gestion et management",
    '1.shs.hist'   : "Histoire",
    '1.shs.infocom': "Sciences de l'information et de la communication",
    '1.shs.ling'   : "Linguistique",
    '1.shs.litt'   : "Littératures",
    '1.shs.phil'   : "Philosophie",
    '1.shs.psy'    : "Psychologie",
    '1.shs.scipo'  : "Science politique",
    '1.shs.socio'  : "Sociologie",
    'NONE'         : "Pas de discipline",
}


def sweet(dom):
    return SWEET[dom]


def header(ws, vals):
    res = []
    for v in vals:
        res.append(cell(ws, v, True))
    ws.append(res)
    return res


def cell(ws, value, bold=False, background=None, color='00000000', comment=None):
    c = WriteOnlyCell(ws, value=value)
    if background is not None:
        c.fill = PatternFill(start_color=background, end_color=background, fill_type='solid')
    c.font = Font(name='Calibri', bold=bold, color=color)
    if comment is not None:
        c.comment = Comment(text=comment, author="DG")
    return c
    
    
class OneSegNoun:

    errors = {}
    errors['ERROR_NPP_2_NC']    = 0
    errors['ERROR_ORTHO']        = 0
    errors['ERROR_UNKNOWN_CHAR'] = 0
    errors['ERROR_UNKNOWN_NC']   = 0
    errors['ERROR_UNKNOWN_NPP']  = 0
    errors['ERROR_ADJ']          = 0
    errors['ERROR_E_SEMI']       = 0
    errors['ERROR_S_AS_NOUN']    = 0
    errors['ERROR_ENGLISH_WORD'] = 0
    errors['ERROR_ENGLISH_NOUN'] = 0
    errors['ERROR_PLURAL_NPP']   = 0
    errors['ERROR_NC_2_NPP']     = 0
    errors['ERROR_CONCAT']       = 0
    
    NOUNS   = {}
    DOMAINS = {}

    # None indique qu'il n'y aura pas de recovery (le titre sera droppé)
    # Le second indique le type d'erreur/correction a enregistré dans OneSegNoun.errors
    CORRECTED = {
        '?évènement::NC'        : ('événement::NC', 'ERROR_ORTHO'),
        '?Global::NPP'          : (None, 'ERROR_ADJ'),
        '?indicateus::NC'       : ('?indicateus::NC', 'ERROR_ORTHO'),
        '?ASSUBILITÉ::NC'       : ('assubilité::NC', 'ERROR_UNKNOWN_NC'),
        '?Dupuit::NPP'          : ('Dupuit::NPP', 'ERROR_UNKNOWN_NPP'),
        '?bitcoin::NC'          : ('bitcoin::NPP', 'ERROR_NC_2_NPP'),
        '?Effet::NPP'           : ('effet::NC', 'ERROR_NPP_2_NC'),
        '?Predicting::NPP'      : (None, 'ERROR_ENGLISH_WORD'),
        '?Cognition::NC'        : ('cognition::NC', 'ERROR_UNKNOWN_NC'),
        '?Adolescence::NPP'     : ('adolescence::NC', 'ERROR_NPP_2_NC'),
        '?Ingres::NPP'          : ('Ingres::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Poussin::NPP'         : ('Poussin::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Crétacé::NPP'         : ('Crétacé::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Cardioceratidae::NPP' : ('Cardioceratidae::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Sédimentologie::NPP'  : ('sédimentologie::NC', 'ERROR_NPP_2_NC'),
        '?Hydrogéochimie::NC'   : ('hydrogéochimie::NC', 'ERROR_UNKNOWN_NC'),
        '?Vertébrés::NPP'       : ('vertébré::NC', 'ERROR_NPP_2_NC'),
        '?Nano::NPP'            : (None, 'ERROR_ADJ'),
        '?Spéciation::NC'       : ('spéciation::NC', 'ERROR_UNKNOWN_NC'),
        '?Nanocomposites::NPP'  : ('nanocomposite::NC', 'ERROR_NPP_2_NC'),
        '?Photomatériaux::NPP'  : ('photomatériau::NC', 'ERROR_NPP_2_NC'),
        '?Nanomatériaux::NPP'   : ('nanomatériau::NC', 'ERROR_NPP_2_NC'),
        '?Viscoélasticité::NC'  : ('viscoélasticité::NC', 'ERROR_UNKNOWN_NC'),
        '?Hydroformylation::NC' : ('hydroformulation::NC', 'ERROR_UNKNOWN_NC'),
        '?Hydroconversion::NC'  : ('hydroconversion::NC', 'ERROR_UNKNOWN_NC'),
        '?HREELS::NC'           : ('?HREELS::NC',),
        '?microréacteur::NC'    : ('microréacteur::NC', 'ERROR_UNKNOWN_NC'),
        '?Chimisorption::NC'    : ('chimisorption::NC', 'ERROR_UNKNOWN_NC'),
        '?Nihon::NPP'           : ('?Nihon::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Environmental::NPP'   : (None, 'ERROR_ADJ'),
        '?IVG::NPP'             : ('IVG::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Glossarium::NC'       : ('?Glossarium::NC', 'ERROR_UNKNOWN_NC'),
        '?Via::NPP'             : ('?Via::NPP', 'ERROR_UNKNOWN_NPP'),
        '?sonification::NC'     : ('sonification::NC', 'ERROR_UNKNOWN_NC'),
        '?Ethnobotanique::NC'   : ('ethnobotanique::NC', 'ERROR_UNKNOWN_NC'),
        '?Venises::NPP'         : ('?Venises::NPP', 'ERROR_PLURAL_NPP'),
        '?Cyber::NPP'           : (None, 'ERROR_ADJ'),
        '?Cooccurrences::NC'    : ('cooccurrences::NC', 'ERROR_UNKNOWN_NC'),
        '?Quantiﬁcation::NC'    : ('quantification::NC', 'ERROR_ORTHO'),
        '?Phénoplasticité::NC'  : ('phénoplasticité::NC', 'ERROR_UNKNOWN_NC'),
        '?Pléistoscène::NC'     : ('Pléistoscène::NPP', 'ERROR_UNKNOWN_NPP'),
        '?démotorisation::NC'   : ('démotorisation::NC', 'ERROR_UNKNOWN_NC'),
        '?maritimisation::NC'   : ('maritimisation::NC', 'ERROR_UNKNOWN_NC'),
        '?DemoMed::NPP'         : ('?DemoMed::NPP',),
        '?SIDA::NPP'            : ('SIDA::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Konso::NC'            : ('?Konso::NC',),
        '?Africains::NC'        : ('africain::NC', 'ERROR_UNKNOWN_NC'),
        '?SIRENA::NPP'          : ('SIRENA::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Beachrocks::NPP'      : ('Beachrocks::NPP', 'ERROR_UNKNOWN_NPP'),
        '?NBIC::NPP'            : ('NBIC::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Proust::NPP'          : ('Proust::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Montaigne::NPP'       : ('Montaigne::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Flaubert::NPP'        : ('Flaubert::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Perceforest::NPP'     : ('Perceforest::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Hypermédias::NPP'     : ('hypermédia::NC', 'ERROR_NPP_2_NC'),
        '?Autoformation::NPP'   : ('autoformation::NC', 'ERROR_NPP_2_NC'),
        '?autoformation::NC'    : ('autoformation::NC', 'ERROR_UNKNOWN_NC'),
        '?EAO::NPP'             : ('EAO::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Approche::NPP'        : ('approche::NC', 'ERROR_NPP_2_NC'),
        '?Learning::NPP'        : ('learning::NC', 'ERROR_ENGLISH_NOUN'),
        '?Corbusier::NPP'       : ('Corbusier::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Fréart::NPP'          : ('Fréart::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Spinoza::NPP'         : ('Spinoza::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Bergson::NPP'         : ('Bergson::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Leibniz::NPP'         : ('Leibniz::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Diderot::NPP'         : ('Diderot::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Nietzsche::NPP'       : ('Nietzsche::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Kant::NPP'            : ('Kant::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Foucault::NPP'        : ('Foucault::NPP', 'ERROR_UNKNOWN_NPP'),
        'Henri ?Poincaré::NPP'  : ('Henri Poincaré::NPP', 'ERROR_CONCAT'),
        '?Poincaré::NPP'        : ('Poincaré::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Habermas::NPP'        : ('Habermas::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Marx::NPP'            : ('Marx::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Malebranche::NPP'     : ('Malebranche::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Carnap::NPP'          : ('Carnap::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Einstein::NPP'        : ('Einstein::NPP', 'ERROR_UNKNOWN_NPP'),
        '?memoriam::NC'         : ('memoriam::NC', 'ERROR_UNKNOWN_NC'),
        '?PAC::NPP'             : ('PAC::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Morphogénèse::NPP'    : ('morphogénèse::NC', 'ERROR_NPP_2_NC'),
        '?RSE::NPP'             : ('RSE::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Saillance::NPP'       : ('saillance::NC', 'ERROR_NPP_2_NC'),
        '?Ethnocritique::NPP'   : ('ethnocritique::NC', 'ERROR_NPP_2_NC'),
        '?Twitter::NPP'         : ('Twitter::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Cohomologie::NC'      : ('cohomologie::NC',),
        '?Compactification::NC' : ('compactification::NC', 'ERROR_UNKNOWN_NC'),
        '?Cohomologie::NPP'     : ('cohomologie::NC', 'ERROR_NPP_2_NC'),
        '?Ondelettes::NC'       : ('ondelette::NC', 'ERROR_UNKNOWN_NC'),
        '?Métamodèles::NPP'     : ('métamodèle::NC', 'ERROR_NPP_2_NC'),
        '?Fibrés::NC'           : (None, 'ERROR_ADJ'),
        '?cK¢::NPP'             : ('?cK¢::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Splines::NPP'         : ('spline::NC', 'ERROR_NPP_2_NC'),
        '?Maths::NPP'           : ('maths::NC', 'ERROR_NPP_2_NC'),
        '?Mixmod::NPP'          : ('Mixmod::NPP', 'ERROR_UNKNOWN_NPP'),
        '?MIXMOD::NPP'          : ('Mixmod::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Pancyclisme::NC'      : ('pancyclisme::NC', 'ERROR_NPP_2_NC'),
        '?Monoïdes::NPP'        : ('monoïde::NC', 'ERROR_NPP_2_NC'),
        '?Finitude::NPP'        : ('finitude::NC', 'ERROR_NPP_2_NC'),
        '?cocycle::NC'          : ('cocycle::NC', 'ERROR_UNKNOWN_NC'),
        '?Cohomologie ?MATHFORMULA::NPP' : ('cohomologie::NC', 'ERROR_NPP_2_NC'),
        '?Paramétrisation::NC'  : ('paramétrisation::NC', 'ERROR_UNKNOWN_NC'),
        '?Clustering::NPP'      : ('clustering::NC', 'ERROR_NPP_2_NC'),
        '?Semi::NPP'            : (None, 'ERROR_E_SEMI'),
        '?Jean::NPP'            : ('Jean::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Paris::NPP'           : ('Paris::NPP', 'ERROR_UNKNOWN_NPP'),
        '?OGM::NPP'             : ('OGM::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Morphogenèse::NPP'    : ('morphogenèse::NC', 'ERROR_NPP_2_NC'),
        '?Aristote::NPP'        : ('Aristote::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Pierre ?Bayle::NPP'   : ('Pierre Bayle::NPP', 'ERROR_CONCAT'),
        '?Bernard::NPP'         : ('Bernard::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Proudhon::NPP'        : ('Proudhon::NPP', 'ERROR_UNKNOWN_NPP'),
        '?SIG::NPP'             : ('SIG::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Learning ?region::NPP'  : ('région::NC', 'ERROR_ENGLISH_NOUN'),
        '?Jacques ?Androuet::NPP' : ('Jacques Androuet::NPP', 'ERROR_CONCAT'),
        '?Claude ?Perrault::NPP'  : ('Claude Perrault::NPP', 'ERROR_CONCAT'),
        '?The::NPP'             : (None, 'ERROR_ENGLISH_WORD'),
        '?E::NPP'               : (None, 'ERROR_E_SEMI'),
        '?Corrélats::NC'        : ('corrélat::NC', 'ERROR_UNKNOWN_NC'),
        '?data::NC'             : ('data::NC', 'ERROR_UNKNOWN_NC'),
        '?Nihon ?fuzai::NPP'    : ('?Nihon ?fuzai::NPP', 'ERROR_CONCAT'),
        '?care::NC'             : ('care::NC', 'ERROR_UNKNOWN_NC'), 
        '?Glossarium::NC'       : ('?Glossarium::NC', 'ERROR_UNKNOWN_NC'),
        '?ADN::NC'              : ('ADN::NC', 'ERROR_UNKNOWN_NC'),
        '?Via Romana::NPP'      : ('Via Romana::NPP', 'ERROR_CONCAT'),
        '?Venises::NPP'         : ('Venise::NPP', 'ERROR_PLURAL_NPP'),
        '?New ?Urbanism::NPP'   : ('urbanisme::NC', 'ERROR_ENGLISH_NOUN'),
        '?DemoMed::NPP'         : ('?DemoMed::NPP',),
        '?Konso::NC'            : ('?Konso::NC',),
        '?Synthesis ?of::NPP'   : ('synthèse::NC', 'ERROR_ENGLISH_NOUN'),
        '?HREELS::NC'           : ('?HREELS::NC',),
        '?rétacé::NPP'          : ('Crétacé::NPP', 'ERROR_UNKNOWN_CHAR'),
        'crétacé::NC'           : ('Crétacé::NPP', 'ERROR_NC_2_NPP'),
        '?œuvre::NC'            : ('oeuvre::NC', 'ERROR_UNKNOWN_CHAR'),
        '?Jean ?Cocteau::NPP'   : ('Jean Cocteau::NPP', 'ERROR_CONCAT'),
        '?Freud::NPP'           : ('Freud::NPP', 'ERROR_UNKNOWN_NPP'),
        '?Etat::NPP'            : ('État::NC',),
        '?Global ?Risk::NPP'    : ('risque::NC', 'ERROR_ENGLISH_NOUN'),
        '?indicateus::NC'       : ('indicateur::NC', 'ERROR_ORTHO'),
        '?Synthčse::NPP'        : ('synthèse::NC', 'ERROR_ORTHO'),
        '?carbènes::NC'         : ('carbène::NC', 'ERROR_UNKNOWN_NC'),
        '?sol::NC'              : ('sol::NC', 'ERROR_UNKNOWN_NC'),
        '?Treatment::NC'        : ('traitement::NC', 'ERROR_ENGLISH_NOUN'),
        'Teneur::NPP'           : ('teneur::NC', 'ERROR_NPP_2_NC'),
        '?Polyhandicap::NPP'    : ('polyhandicap::NC', 'ERROR_NPP_2_NC'),
    }

    
    @classmethod
    def select(cls, **filters):
        res = []
        for k, n in cls.NOUNS.items():
            if n.nb_head['all'] < 10: continue
            ok = True
            for fk, fv in filters.items():
                if getattr(n, fk) < fv:
                    ok = False
                if ok:
                    res.append(n)
        if 'med_dom' in filters:
            return sorted(res, key=lambda e: e.med_dom, reverse=True)
        else:
            return res
    
    
    @classmethod
    def lex(cls, data, segnum=0, title=None, calc=True):
        if segnum == 2: # do both for 2 seg titles
            print('>>> LEX both')
            cls.lex(data, segnum=0, calc=False)
            cls.lex(data, segnum=1, calc=False)
        elif segnum == 'couple': # do by couple head1--head2 for 2 seg titles
            print('>>> LEX couple')
            for kt, t in data.items():
                if t.domain in ['NONE', '1.shs.autre']: continue
                if t.nb_segments != 2: # t.segments contient aussi des seg non segmentant ! (. finaux)
                #if len(t.segments) != 1 or t.nb_segments != 2:
                    print(t)
                    print('Len of segments: ', len(t.segments))
                    print('Val of nb_segments: ', t.nb_segments)
                    t.info()
                    raise Exception('Not a two segment title!')
                head1 = t.words[t.roots[0]]
                head2 = t.words[t.roots[1]]
                cls.record(head1.lemma + '--' + head2.lemma, head1.pos + '--' + head2.pos, True, t.domain)
        else: # do for segnum in [0,1]
            if calc: print('>>> LEX ' + str(segnum))
            for kt, t in data.items():
                if t.domain in ['NONE', '1.shs.autre']: continue
                for cpt, word in enumerate(t.words):
                    # if segnum=0, we do only the first segment!
                    if segnum == 0 and len(t.segments) > 0 and cpt == t.segments[0]: break
                    # if segnum=1, we do nothing in the first segment!
                    elif segnum == 1 and len(t.segments) > 0 and cpt < t.segments[0]: continue
                    #elif segnum == 1 and len(t.segments) == 0: raise Exception('Title is not bisegmental')
                    # dans le cas où on fait ça pour tout, on va passer sur des titres monoseg avec segnum = 1
                    # donc si on veut enquêter sur le second seg (segnum = 1) alors qu'on a nb_segments == 1
                    # on quitte le titre
                    elif segnum == 1 and t.nb_segments == 1: break
                    if word.pos in ['NC', 'NPP']:
                        head = (cpt == t.roots[segnum]) # fetch the root corresponding to segnum
                        if cpt + 1 < len(t.words) and t.words[cpt + 1].pos == 'NPP' and word.pos == 'NPP':
                            lemma = word.lemma + ' ' + t.words[cpt + 1].lemma # 'André Breton'
                        elif cpt + 2 < len(t.words) and t.words[cpt + 1].lemma == 'de' and t.words[cpt + 2].lemma == 'NPP':
                            lemma = word.lemma + ' de ' + t.words[cpt + 2].lemma # Bernard de Clairvaux
                        elif cpt > 0 and word.pos == 'NPP' and t.words[cpt - 1].lemma == 'saint':
                            lemma = t.words[cpt - 1].lemma + ' ' + word.lemma # Saint Bernard
                        elif cpt + 2 < len(t.words) and word.lemma in ['?E', '?e'] and t.words[cpt + 1].lemma == '-': #and ('E-' in t.text or 'e-' in t.text):
                            lemma = 'e-' + t.words[cpt + 2].lemma # e-management
                            cls.errors['ERROR_E_SEMI'] += 1
                        elif cpt + 2 < len(t.words) and word.lemma in ['?Semi', '?semi'] and t.words[cpt + 1].lemma == '-': #and ('Semi-' in t.text or 'semi-' in t.text):
                            lemma = 'semi-' + t.words[cpt + 2].lemma # semi-figement
                            cls.errors['ERROR_E_SEMI'] += 1
                        elif word.lemma == '?s':
                            if cpt > 1 and t.words[cpt-1].lemma in ['(', '.'] and t.words[cpt-2].pos in ['NC', 'NPP']:
                                lemma = t.words[cpt - 2].lemma # *mobilité*(_s_), *mobilité*._s_
                                cls.errors['ERROR_S_AS_NOUN'] += 1
                            elif cpt + 2 < len(t.words) and t.words[cpt+1].lemma == ')' and t.words[cpt+2].pos in ['NC', 'NPP']:
                                lemma = t.words[cpt + 2].lemma # Quelle(_s_) *diversité*(s)
                                cls.errors['ERROR_S_AS_NOUN'] += 1
                        elif word.lemma == 'mise':
                            if cpt + 2 < len(t.words):
                                if t.words[cpt + 2].lemma == '?œuvre':
                                    lemma = 'mise ' + t.words[cpt + 1].lemma + ' oeuvre'
                                else:
                                    lemma = 'mise ' + t.words[cpt + 1].lemma + ' ' + t.words[cpt + 2].lemma
                            else:
                                print('mise change:', t)
                        else:
                            lemma = word.lemma
                        cls.record(lemma, word.pos, head, t.domain)
        if calc: # MUST BE CALLED ONLY ONCE! OR HEADS of DOM are counted twice!
            num = segnum if segnum == 2 else 1
            cls.calc(data, num_by_title=num)
        if title is not None:
            cls.to_sheet(title)

    # unused
    @classmethod
    def reset(cls):
        cls.NOUNS   = {}
        cls.DOMAINS = {}
        cls.ALREADY_CALC = False
        cls.errors['ERROR_NPP_2_NC']     = 0
        cls.errors['ERROR_ORTHO']        = 0
        cls.errors['ERROR_UNKNOWN_CHAR'] = 0
        cls.errors['ERROR_UNKNOWN_NC']   = 0
        cls.errors['ERROR_UNKNOWN_NPP']  = 0
        cls.errors['ERROR_ADJ']          = 0
        cls.errors['ERROR_E_SEMI']       = 0
        cls.errors['ERROR_S_AS_NOUN']    = 0
        cls.errors['ERROR_ENGLISH_WORD'] = 0
        cls.errors['ERROR_ENGLISH_NOUN'] = 0
        cls.errors['ERROR_PLURAL_NPP']   = 0
        cls.errors['ERROR_NC_2_NPP']     = 0
        cls.errors['ERROR_CONCAT']       = 0
    
    ALREADY_CALC = False
    SEUIL_FREQ = 0.003 #0.001 #
    SEUIL_OCC  = 0.025

    TOTAL_DOM      = 0
    TOTAL_HEAD_OCC = 0
    TOTAL_HEAD_LEM = 0
    
    @classmethod
    def calc(cls, data, num_by_title=1):
        print('>>> CALC')
        if cls.ALREADY_CALC: raise Exception("Calc must be called only once!")
        else: cls.ALREADY_CALC = True
        NBDOM = len(OneSegNoun.DOMAINS)
        EGAL_DISTRI = 1 / NBDOM
        for kd, d in cls.DOMAINS.items():
            d.count()
        for k, n in cls.NOUNS.items():
            n.nb_doms = 0 # len(n.nb_head)-1 # ATTENTION APRES len(n.nb_head)-1 == NBDOM et pas NBDOM où la tête est présente !
            n.egal_distri = n.nb_head['all'] / NBDOM
            n.chi2 = 0
            # Calcul de la moyenne des fréquence dans les domaines
            n.freq_dom = {}
            n.freq_lem = {}
            n.disc_dom = []
            for kd, d in cls.DOMAINS.items():
                if kd == 'all': continue
                if kd == 'NONE': raise Exception('Should have a domain')
                if kd not in n.nb_head: # ON AJOUTE LES DOMAINES OU IL N'Y A PAS LA TETE
                    n.nb_head[kd] = 0
                if n.nb_head[kd] > 0:
                    n.nb_doms += 1
                n.freq_dom[kd] = n.nb_head[kd] / d.total_heads
                if n.nb_head['all'] == 0:
                    n.freq_lem[kd] = 0
                else:
                    n.freq_lem[kd] = n.nb_head[kd] / n.nb_head['all']
                    n.chi2 += ((n.nb_head[kd] - n.egal_distri)**2) / n.egal_distri
                # Disc Calc
                if d.total_heads * OneSegNoun.SEUIL_FREQ >= 1:
                    if n.freq_dom[kd] >= OneSegNoun.SEUIL_FREQ:
                        d.filter1 += 1    
                        if n.freq_lem[kd] - EGAL_DISTRI >= OneSegNoun.SEUIL_OCC:
                            d.filter2 += 1
                            n.disc_dom.append(kd)
                # End Disc Calc
            n.moy_dom = fmoy(n.freq_dom.values())
            n.ect_dom = fect(n.freq_dom.values())
            n.med_dom = fmed(n.freq_dom.values())
            n.in_tutin = 1 if n.lemma in TUTIN else 0
        # On calcul le nombre total de lemmes différents de têtes pris en compte
        cls.TOTAL_HEAD_LEM = 0
        for _, n in cls.NOUNS.items():
            for kd in n.nb_head:
                if n.nb_head[kd] != 0:
                    cls.TOTAL_HEAD_LEM += 1
                    break
        # On calcul le nombre total d'occurrences de têtes
        cls.TOTAL_HEAD_OCC = 0
        for kd, d in cls.DOMAINS.items():
            cls.TOTAL_HEAD_OCC += d.total_heads
        # On calcul le nombre de domaine pris en compte
        cls.TOTAL_DOM = len(cls.DOMAINS)

    
    @classmethod
    def correct(cls, key):
        if key in cls.CORRECTED:
            res = cls.CORRECTED[key]
            if type(res) != tuple:
                print(key, res)
            key = res[0]
            if len(res) > 1:
                msg = res[1]
                cls.errors[msg] += 1
                if key is None: # msg + non recovery
                    return None, None, None
        lemma, pos = key.split('::')
        return key, lemma, pos

    
    @classmethod
    def record(cls, lemma, pos, head, domain):
        # recording noun
        key = lemma + '::' + pos
        key, lemma, pos = cls.correct(key)
        if key is None: return # preventing registering adj
        if key not in cls.NOUNS:
            n = OneSegNoun(lemma, pos)
            cls.NOUNS[key] = n
        else:
            n = cls.NOUNS[key]
        n.nb_occ['all'] += 1
        if domain in n.nb_occ:
            n.nb_occ[domain] += 1
        else:
            n.nb_occ[domain] = 1
        if head:
            n.nb_head['all'] += 1
            if domain in n.nb_head:
                n.nb_head[domain] += 1
            else:
                n.nb_head[domain] = 1
        # recording domain
        if domain not in cls.DOMAINS:
            d = Domain(domain)
            cls.DOMAINS[domain] = d
        else:
            d = cls.DOMAINS[domain]
        d.record(key, head)

    
    def __init__(self, lemma, pos):
        self.lemma = lemma
        self.pos = pos
        self.nb_occ = { 'all' : 0 }
        self.nb_head = { 'all' : 0 }
    

    def __str__(self):
        star = '' if self.lemma in TUTIN else '*'
        return self.lemma + star

    
    def __repr__(self):
        return str(self)


    DOMAINS_BAG_OF_WORDS = {}
    
    @classmethod
    def disciplinary_neo(cls):
        def tfidf(osn, kdom):
            tf = osn.freq_dom[kdom]
            idf = math.log10(25 / osn.nb_doms)
            return tf * idf
        # Calculate TF*IDF for each heads of each domains
        # '0.math' : ['discipline::NC', 'quantification::NC']
        all_res = {}
        # 'discipline::NC' : ['0.math', '0.info', ...]
        all_lem = {}
        for kd in cls.DOMAINS:
            res = {}
            for k, v in cls.DOMAINS[kd].heads.items():
                if v == 0: continue
                osn = cls.NOUNS[k]
                res[k] = tfidf(osn, kd)
                if k not in all_lem:
                    all_lem[k] = [kd]
                else:
                    all_lem[k].append(kd)
            all_res[kd] = sorted(res, key=res.get, reverse=True)
        # Output
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet('Disciplines')
        i = 1
        # Title page
        nb_titles = stat('domain', display=False)
        header(ws, ['N°', 'Discipline', 'Code', 'Head Lem diff', 'Head Occ diff', 'Nb Titles'])
        for kd in all_res:  # len(all_res[kd]) == cls.DOMAINS[kd].nb_heads
            ws.append([i, sweet(kd), kd, len(all_res[kd]), cls.DOMAINS[kd].total_heads, nb_titles[(kd,)]])
            i += 1
        # All lemma page
        ws = wb.create_sheet('Lemma')
        header(ws, ['Lemma', 'POS', 'Nb Dom'])
        for kl in all_lem:
            row = kl.split('::')
            row.append(len(all_lem[kl]))
            osn = OneSegNoun.NOUNS[kl]
            for kd in all_lem[kl]:
                row.append(kd)
                row.append(tfidf(osn, kd))
            ws.append(row)
        # All domains as a bag of words
        ws = wb.create_sheet('Domains')
        row = ['Lemme', 'POS']
        for kd in all_res:
            row.append(kd)
            cls.DOMAINS_BAG_OF_WORDS[kd] = []
        header(ws, row)
        for kl in all_lem:
            row = kl.split('::')
            osn = OneSegNoun.NOUNS[kl]
            for kd in all_res:
                if kd in all_lem[kl]:
                    v = tfidf(osn, kd)
                else:
                    v = 0
                row.append(v)
                cls.DOMAINS_BAG_OF_WORDS[kd].append(v)
            ws.append(row)
        # Distance between domains page
        ws = wb.create_sheet('Distances')
        row = ['Distances']
        for kd in cls.DOMAINS_BAG_OF_WORDS:
            row.append(kd)
        header(ws, row)
        for kd1 in cls.DOMAINS_BAG_OF_WORDS:
            row = [kd1]
            maxx = len(cls.DOMAINS_BAG_OF_WORDS[kd1]) # always the same, square matrix
            for kd2 in cls.DOMAINS_BAG_OF_WORDS:
                tt = 0
                for i in range(0, maxx):
                    tt += (cls.DOMAINS_BAG_OF_WORDS[kd1][i] - cls.DOMAINS_BAG_OF_WORDS[kd2][i]) ** 2
                tt = math.sqrt(tt)
                row.append(tt)
            ws.append(row)
        # Domain pages
        cpt_d = 1
        for kd in all_res:
            ws = wb.create_sheet(str(cpt_d) + '. ' + kd)
            i = 1
            header(ws, ['N°', 'Lemma', 'POS', 'Count', 'TF*IDF', 'TF', 'IDF', 'Nb dom'])
            for k in all_res[kd]:
                osn = OneSegNoun.NOUNS[k]
                v = tfidf(osn, kd)
                ws.append([i] + k.split('::') + [osn.nb_head[kd], v, osn.freq_dom[kd], math.log10(25 / osn.nb_doms), osn.nb_doms])
                i += 1
            cpt_d += 1
        wb.save('blob.xlsx')


    @classmethod
    def disciplinary(cls, display_limit=None):
        if not cls.ALREADY_CALC: raise Exception("Calc must be called before!")
        disc_domains = {}
        all_filtered = []
        output_lines = []
        TOTAL = 0
        for kd, d in cls.DOMAINS.items():
            if d.total_heads * OneSegNoun.SEUIL_FREQ < 1: continue
            if kd == 'NONE': raise Exception('Should have a domain')
            filtered = []
            for kn, n in cls.NOUNS.items():
                if kd in n.disc_dom:
                    filtered.append(n)
                    TOTAL += 1
                    if n not in all_filtered:
                        all_filtered.append(n)
            output_lines.append(f'{sweet(kd)}\n')
            output_lines.append(f'Filter1 : {d.filter1:6d} heads / {d.nb_heads:6d} ({round(d.filter1 / d.nb_heads * 100, 2)} %)\n')
            output_lines.append(f'Filter2 : {d.filter2:6d} heads / {d.nb_heads:6d} ({round(d.filter2 / d.nb_heads * 100, 2)} %)\n')
            res = sorted(filtered, key=lambda e: e.freq_dom[kd], reverse=True)
            disc_domains[kd] = []
            unknown = []
            output_lines.append("    lemma             POS    F / dom F / o  occ [ /   dom  /   occ\n")
            cpt = 0
            for n in res:
                # lemma, pos
                # freq/total head (la fréq rel dans ce dom)
                # freq/nb occ total (par rapport au nb d'occ de cette tête)
                # nb occ dans ce dom
                # total head dans de dom
                # nb occ total
                if display_limit is not None and ((type(display_limit) == float and n.freq_dom[kd] < display_limit) or (type(display_limit) == int and cpt == display_limit)): break
                cpt += 1
                output_lines.append(f"    {n.lemma:18} {n.pos:4} {n.freq_dom[kd]:7.6f} {n.nb_head[kd] / n.nb_head['all']:4.3f} {n.nb_head[kd]:4d} [ / {d.total_heads:5d}  / {n.nb_head['all']:5d} ] {len(n.disc_dom)}\n")
                disc_domains[kd].append(n)
                if n.lemma[0] == '?':
                    unknown.append(n)
            if len(unknown) > 0:
                output_lines.append('------- Unknown -------\n')
                for u in unknown:
                    output_lines.append(f"{u.lemma} {u.pos} {u.nb_head[kd]}\n")
        output_lines.append(f'Total Selected : {len(all_filtered)}')
        # Output to File
        filename = f'heads_disc_Fr_{OneSegNoun.SEUIL_FREQ}_Oc_{OneSegNoun.SEUIL_OCC}_NbH_{len(all_filtered)}_TotDD_{TOTAL}'
        f = open(filename + '.txt', mode='w', encoding='utf8')
        for line in output_lines:
            f.write(line)
        f.close()
        # Output to Excel
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet('Disciplinary')
        ws.append(['Dom', 'Kept A', 'Nb heads', 'Total heads', 'Sel', '%', 'Kept B', '%'])
        for kd in sorted(SWEET, key=SWEET.get):
            if kd in ['NONE', '1.shs.autre']: continue
            d = cls.DOMAINS[kd]
            if d.total_heads * OneSegNoun.SEUIL_FREQ < 1: continue
            ws.append([sweet(kd), d.filter2, d.nb_heads, d.total_heads, d.filter1, d.filter1 / d.nb_heads, d.filter2, d.filter2 / d.nb_heads])
        # Sheet for Lemma
        ws = wb.create_sheet('By Lemma')
        ws.append(
            [cell(ws, value='Lemma', bold=True, background='00D3D3D3'),
             cell(ws, value='POS', bold=True, background='00D3D3D3'),
             cell(ws, value='Nb Dom', bold=True, background='00D3D3D3'),
             cell(ws, value='Disc Dom', bold=True, background='00D3D3D3'),
             cell(ws, value='%', bold=True, background='00D3D3D3')
            ])
        for n in all_filtered:
            row = [n.lemma, n.pos, n.nb_doms, len(n.disc_dom), len(n.disc_dom) /  n.nb_doms]
            for kd in n.disc_dom:
                row.append(sweet(kd))
            ws.append(row)
        # Sheet for Domain
        ws = wb.create_sheet('By Domain Order by freq_dom')
        ws.append([cell(ws, value='Domain', bold=True, background='00D3D3D3'), '1', '2', '3', '4', '5', '6', '7', '8', '9', '10'])
        for kd, nouns in disc_domains.items():
            row = [sweet(kd)]
            for n in sorted(nouns, key=lambda e: e.freq_dom[kd], reverse=True):
                row.append(n.lemma + ' ' + n.pos)
            ws.append(row)
        # Sheet for Domain
        ws = wb.create_sheet('By Domain Order by freq_lem')
        ws.append([cell(ws, value='Domain', bold=True, background='00D3D3D3'), '1', '2', '3', '4', '5', '6', '7', '8', '9', '10'])
        for kd, nouns in disc_domains.items():
            row = [sweet(kd)]
            for n in sorted(nouns, key=lambda e: e.freq_lem[kd], reverse=True):
                row.append(n.lemma + ' ' + n.pos)
            ws.append(row)
        wb.save(filename + '.xlsx')

    
    @classmethod
    def to_sheet(cls, title):
        if not cls.ALREADY_CALC: raise Exception("Calc must be called before!")
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet('1seg head')
        # Title
        nbdom = ' (' + str(cls.TOTAL_DOM) + ')'
        title_row = ['LEMMA', 'POS', 'Tutin', 'NB HEAD', '%', 'NB OCC', '%', 'HEAD/OCC', 'NBDOM-HD' + nbdom,
                     '%NBDOM-HD', 'Moy DOM', 'Med DOM', 'ECT DOM', 'NBDOM OCC' + nbdom]
        for i in range(0, cls.TOTAL_DOM):
            title_row.extend(['Dom', 'nb', 'f/dom', 'f/lem'])
        ws.append(title_row)
        # Data
        for k, n in cls.NOUNS.items():
            nb_occ = n.nb_occ['all']
            nb_head = n.nb_head['all']
            if fmax(n.freq_dom.values()) < OneSegNoun.SEUIL_FREQ: continue
            row = [n.lemma,                         # Lemma
                   n.pos,                           # Pos
                   n.in_tutin,                      # In Tutin
                   nb_head,                         # NB HEAD
                   nb_head / cls.TOTAL_HEAD_OCC,    # % de toutes les heads
                   nb_occ,                          # NB OCC
                   nb_occ / Domain.total_nouns,     # % de toutes les occurrences
                   nb_head / nb_occ,                # rapport HEAD / OCC
                   n.nb_doms,                       # NB Dom où la tête est présente, remove 'all'
                   n.nb_doms / cls.TOTAL_DOM,       # % du NB Dom où tête présente / NB Dom
                   n.moy_dom,                       # Fréquence moyenne pour heads du dom
                   n.med_dom,                       # Médiane
                   n.ect_dom,                       # écart-type des freqs par rapport aux heads du dom
                   len(n.nb_occ)-1,                 # NB Dom où l'occ est présente, remove 'all'
                   ]
            for kd in sorted(n.freq_dom, key=n.freq_dom.get, reverse=True):
                if kd == 'all': continue
                dom = cls.DOMAINS[kd]
                if n.nb_head[kd] == 0: break
                row.extend([
                    kd,
                    n.nb_head[kd],
                    n.freq_dom[kd],
                    n.freq_lem[kd]])
            ws.append(row)
        # Second Sheet
        ws = wb.create_sheet('Lem by dom')
        title_row = ['Lemma', 'POS', 'Nb', 'Egal/' + str(cls.TOTAL_DOM), 'Chi2']
        for kd in cls.DOMAINS:
            title_row.append(kd)
        ws.append(title_row)
        for k, n in cls.NOUNS.items():
            if fmax(n.freq_dom.values()) < OneSegNoun.SEUIL_FREQ: continue
            row = [n.lemma, n.pos, n.nb_head['all'], n.egal_distri, n.chi2]
            for kd in cls.DOMAINS:
                row.append(n.nb_head[kd])
            ws.append(row)
        wb.save(title)
        return
        # DEPRECATED
        # Second Sheet
        FIRST_X = 20
        ws = wb.create_sheet('1seg best ' + str(FIRST_X) + ' by dom')
        title_row = []
        lemma = {} # 'Brève' => [0.qfin, 0.info]
        for kd, d in cls.DOMAINS.items():
            title_row.extend([kd, 'nb', '%'])
        ws.append(title_row)
        for i in range(0, FIRST_X):
            row = []
            for kd, d in cls.DOMAINS.items():
                if i < len(d.best_heads):
                    best = d.best_heads[i]
                    row.extend([best, d.heads[best], per(d.heads[best], d.total_heads)])
                    # for next sheet
                    if best not in lemma:
                        lemma[best] = [kd]
                    else:
                        lemma[best].append(kd)
                else:
                    row.extend(['', '', ''])
            ws.append(row)
        # Third Sheet
        ws = wb.create_sheet('1seg best ' + str(FIRST_X) + ' by lemma')
        ws.append(['LEMMA', 'POS', 'NB DOM', 'DOMAINS / NUM...'])
        lemma_count = {}
        for lem in lemma:
            lemma_count[lem] = len(lemma[lem])
        for lem in sorted(lemma_count, key=lemma_count.get, reverse=True):
            row = lem.split('::') # lemma, pos
            row.append(lemma_count[lem]) # number of domain where it is present in the X first
            for kd in lemma[lem]:
                row.extend([kd, cls.DOMAINS[kd].best_heads.index(lem)]) # domains where it is present in the X first + pos
            ws.append(row)
        wb.save(title)


# Lexique des têtes
def make_lexique():
    lexique_1s = {} # Lexique of heads for one segment only titles (1 head by segment)
    lexique_2s = {} # Lexique of heads for two segment titles (2 heads by segment)
    lexique_2s1 = {} # Lexique of heads for the first segment of two segment titles
    lexique_2s2 = {} # Lexique of heads for the second segment of two segment titles
    lexique_allheads = {} # Lexique of all heads
    lexique_allnouns = {} # Lexique of all nouns
    for kt, t in c1n.items():
        root = t.words[t.roots[0]]
        key = root.lemma + '::' + root.pos
        record(key, lexique_1s)
        record(key, lexique_allheads)
        for word in t.words:
            if word.pos in ['NC', 'NPP']:
                key = word.lemma + '::' + word.pos
                record(key, lexique_allnouns)
    for kt, t in c2n.items():
        root1 = t.words[t.roots[0]]
        root2 = t.words[t.roots[1]]
        key1 = root1.lemma + '::' + root1.pos
        key2 = root2.lemma + '::' + root2.pos
        record(key1, lexique_2s)
        record(key2, lexique_2s)
        record(key1, lexique_2s1)
        record(key2, lexique_2s2)
        record(key1, lexique_allheads)
        record(key2, lexique_allheads)
    percent(lexique_1s)
    percent(lexique_2s)
    # I want to know how a word is balanced between first and second seg
    percent(lexique_2s, ref=lexique_2s1, revert=True)
    percent(lexique_2s, ref=lexique_2s2, revert=True)
    percent(lexique_2s1)
    percent(lexique_2s2)
    percent(lexique_allheads)
    # I want to know if a word is more than often a head
    percent(lexique_allheads, ref=lexique_allnouns)
    percent(lexique_allnouns)
    wb = openpyxl.Workbook(write_only=True)
    dump2sheet("1 SEG", lexique_1s, wb)
    dump2sheet("2 SEG", lexique_2s, wb)
    dump2sheet("2 SEG ONE", lexique_2s1, wb)
    dump2sheet("2 SEG TWO", lexique_2s2, wb)
    dump2sheet("ALL HEADS", lexique_allheads, wb)
    dump2sheet("ALL NOUNS", lexique_allnouns, wb)
    wb.save("output.xlsx")


def stats(data=None):
    if data is None: data = titles
    key = 'len_without_ponct'
    dd = Data({ 'len_without_ponct':'num' })
    for kt, t in data.items():
        dd.count(getattr(t, key))
    return dd


# Global stat calulation, deprecated in favor of stat
def calc_stats(titles):
    global stats
    # Basic title stats
    keys = ['nb_restarts', 'nb_segments', 'nb_roots', 'nb', 'year', 'domain', 'typ',
            ('nb_parts', 'nb_segments'), ('domain', 'nb_segments') ]
    for key in keys:
        stat(key)
    # Word stats
    stats['root_pos']   = Data({ 'root_pos':'str', 'count':'num' }) # pos of root
    stats['root_lemma'] = Data({ 'root_lemma':'str', 'pos':'str', 'count':'num', 'is_sgn':'bool' }) # lemma of root
    stats['seg_combi']  = Data({ 'seg_combi':'tuple', 'count':'num' }) # combinaison of seg
    stats['seg_lemma']  = Data({ 'seg_lemma':'tuple', 'count':'num' }) # lemma of seg ponct
    # seg2 dependant from root of seg1
    stats['seg2_nb_dep_from_root'] = Data({ 'seg2_nb_dep_from_root' : 'num', 'count' : 'num' })
    stats['seg2_pos_dep_from_root'] = Data({ 'seg2_pos_dep_from_root' : 'str', 'count' : 'num' })
    stats['seg2_dep_dep_from_root'] = Data({ 'seg2_dep_dep_from_root' : 'str', 'count' : 'num' })
    stats['seg2_lem_dep_from_root'] = Data({ 'seg2_lem_dep_from_root' : 'str', 'pos' : 'str', 'count' : 'num', 'is_sgn' : 'bool' })
    for kt, t in titles.items():
        combi = []
        num_seg = 1
        pos_root = None
        nb_dep_from_root = 0
        for i, w in enumerate(t.words):
            if is_root(w):
                stats['root_pos'].count(w.pos)
                stats['root_lemma'].count(w.lemma + '::' + w.pos)
                pos_root = i + 1 # TALISMANE starts at 1
            if is_seg(w):
                stats['seg_lemma'].count(w.lemma)
                combi.append(w.form)
                num_seg += 1
            if num_seg > 1 and w.gov == pos_root:
                nb_dep_from_root += 1
                stats['seg2_pos_dep_from_root'].count(w.pos)
                stats['seg2_dep_dep_from_root'].count(w.dep)
                stats['seg2_lem_dep_from_root'].count(w.lemma + '::' + w.pos)
        if num_seg > 1:
            stats['seg2_nb_dep_from_root'].count(nb_dep_from_root)
        stats['seg_combi'].count(tuple(combi))
    for k in stats['root_lemma']:
        elements = k.split('::')
        stats['root_lemma'][k]['root_lemma'] = elements[0]
        stats['root_lemma'][k]['pos'] = elements[1]
        stats['root_lemma'][k]['is_sgn'] = is_top_100_signoun(elements[0])
    for k in stats['seg2_lem_dep_from_root']:
        elements = k.split('::')
        stats['seg2_lem_dep_from_root'][k]['seg2_lem_dep_from_root'] = elements[0]
        stats['seg2_lem_dep_from_root'][k]['pos'] = elements[1]
        stats['seg2_lem_dep_from_root'][k]['is_sgn'] = is_top_100_signoun(elements[0])

#-------------------------------------------------
# Boot
#-------------------------------------------------

old    = None
titles = {}
t1     = None # t with 1 part
t11    = None # t with 1 part, 1 segment
t111   = None # t with 1 part, 1 segment, 1 root
t111n  = None # t with 1 part, 1 segment, 1 root of pos in NC, NPP
t112   = None # t with 1 part, 1 segment, 2 roots
t12    = None # t with 1 part, 2 segments
t121   = None # t with 1 part, 2 segments, 1 root
t122   = None # t with 1 part, 2 segments, 2 roots
t2     = None # t with 2 parts
t22    = None # t with 2 parts, 2 segments
t222   = None # t with 2 parts, 2 segments, 2 roots
corpus = None # Working corpus but not filtered on root nature
final  = None # Working corpus final
c1s    = None # Corpus titles with only 1 segment
c2s    = None # Corpus titles with 2 segments
c1n    = None # Corpus titles with only 1 segment root is NC or NPP
c2n    = None # Corpus titles with 2 segments one root is NC,NPP the other in NOUN, VERB, PREP
#cas    = None # Corpus titles with All Segment separeted

disc = {}
disc_flat = []
trans = {}

fast_load           = False
just_load           = True
produce_trans_excel = False
produce_disc_excel  = False
produce_quick       = False
produce_neo_disc    = False

def init(debug):
    global titles, old, t1, t11, t111, t111n, t112, t12, t121, t122, t2, t22, t222, \
           corpus, c1s, c2s, c1n, c2n, final, disc, disc_flat, trans
    load_recoding_table(debug)
    if debug: print('[INFO] --- Domain recode dictionary loaded\n')
    titles = read_titles_metadata(r'data\total-articles-HAL.tsv')
    if fast_load:
        files = [r"data\output_tal_01.txt"]
    else:
        files = [r"data\output_tal_01.txt",
             r"data\output_tal_02.txt",
             r"data\output_tal_03.txt",
             r"data\output_tal_04.txt",
             r"data\output_tal_05.txt",
             r"data\output_tal_06.txt"]
    for file in files:
        read_update_from_talismane_data(titles, file)
    if debug:
        print("[INFO] --- Total Titles :", len(titles))
    # Old info on part & segment ON THE BASIS MATERIAL
    print()
    print("[INFO] --- 2 parts in base data    :", count({'nb_parts' : 2}))
    print("[INFO] --- 2 segments in base data :", count({'nb_segments' : 2}))
    stat('nb_parts')
    print()
    stat('nb_segments')
    print()
    # Selectors
    old = titles
    # 1 part
    #t1 = select({'nb_parts' : 1})           # len = 295 331
    t11 = select({'parts_segments' : '1:1'}) # len = 171 890
    t12 = select({'parts_segments' : '1:2'}) # len =  98 931
    titles = t11
    t111 = select({'nb_roots' : 1})          # len = 171 890
    t112 = select({'nb_roots' : 2})          # len =  20 160
    titles = t12
    t121 = select({'nb_roots' : 1})          # len =  56 851
    t122 = select({'nb_roots' : 2})          # len =  33 631
    titles = t111
    t111n = select({'roots.0.pos' : ['NC', 'NPP']}) # len = 136707
    del t11
    del t12
    # 2 parts : on ne prend que les 2 parts ayant 2 segments et dont la partition se fait sur un segmentateur
    titles = old
    t2x = select({'nb_parts' : 2})           # len =  38 808
    t2 = {}                                  # len =  38 346 (minus 462)
    t22 = {}                                 # len =  30 055
    for k, v in t2x.items():
        if is_seg(v.words[v.restarts[0] - 1]):
            t2[k] = v
            if v.nb_segments == 2:
                t22[k] = v
    del t2x
    titles = t22
    t222 = select({'nb_roots' : 2})
    del t22
    # Info
    titles = old
    print()
    #print('t1 is available :', len(t1))
    print('t111 is available :', len(t111))
    print('t112 is available :', len(t112))
    print('t121 is available :', len(t121))
    print('t122 is available :', len(t122))
    print('t222 is available :', len(t222))
    print('Sum :              ', len(t111)+len(t112)+len(t121)+len(t122)+len(t222))
    #print('t2 is available :', len(t2))
    #print('t22 is available :', len(t22))
    #print('t111 is available :', len(t111))
    #print('t111n is available :', len(t111n))
    # Transform 121 titles (1 part, 2 segments, 1 root, 1:0 or 0:1) into 122 / 1:1 titles
    print('\n* Finding head\n')
    d, i = find_subroot(t121)
    print('Promoted directly by gov by root from other seg:', d)
    print('Promoted by gov from other seg:', i)
    # Gather our final working corpus.
    corpus = {}
    # We take t111.
    corpus.update(t111)
    # We don't take 112, too much problem.
    # We take the updated 121.
    titles = t121
    temp = select({'roots_by_segments' : '1:1'})
    print('Total promoted titles (one subroot found) :', len(temp), '/', len(t121), '(', round(len(temp)/len(t121)*100, 0), '% )')
    corpus.update(temp)
    print()
    titles = old
    # We take 122.
    titles = t122
    temp = select({'roots_by_segments' : '1:1'})
    corpus.update(temp)
    titles = old
    # We take 222.
    titles = t222
    temp = select({'roots_by_segments' : '1:1'})
    corpus.update(temp)
    # We make two subcorpus
    titles = corpus
    c1s = select({'nb_segments' : 1})
    c2s = select({'nb_segments' : 2})
    # Info sur les titres à 2 segments avant triage sur type
    print('\n* Info on 2 seg titles before filtering on type of head\n')
    old = titles
    titles = c2s
    res = stat(['roots.0.pos:agg', 'roots.1.pos:agg'])
    titles=  old
    print()
    # We select title with a NC or NPP root
    titles = c1s
    c1n = select({'roots.0.pos' : ['NC', 'NPP']})
    titles = corpus
    c2n = {}
    for kt, t in c2s.items():
        if (t.words[t.roots[0]].pos in ['NC', 'NPP'] and agg(t.words[t.roots[1]].pos) in ['NOUN', 'VERB', 'PREP']) or \
           (agg(t.words[t.roots[0]].pos) in ['NOUN', 'VERB', 'PREP'] and t.words[t.roots[1]].pos in ['NC', 'NPP']):
                c2n[kt] = t
    #titles = old
    # Final
    final = copy.copy(c1n)
    final.update(c2n)
    titles = final
    print('corpus and final are available')
    print('c1s and c1n are available (one segment)')
    print('c2s and c2n are available (two segments)')
    print('=> n version are with at least one noun root')
    print('Corpus length:', len(corpus), '/', len(titles), '(', p(len(corpus)), '% )')
    print('c1s length:', len(c1s), '(', round(len(c1s)/len(corpus)*100,2), '% )')
    print('c2s length:', len(c2s), '(', round(len(c2s)/len(corpus)*100,2), '% )')
    print()
    print('Set titles to one of these values to change the corpus requested.')
    print('Set titles to old to reset to ALL titles')
    print('NB titles is set to final')
    #make_lexique()

    #--------------------
    # Production
    #--------------------
    
    #OneSegNoun.lex(c1n, 0)
    #output_filename = "heads_c1n.xlsx"
    #OneSegNoun.lex(c2n, 0)
    #output_filename = "heads_c2nA.xlsx"
    #OneSegNoun.lex(c2n, 1)
    #output_filename = "heads_c2nB.xlsx"
    OneSegNoun.lex(final, 2)
    #output_filename = "heads_corpus.xlsx"
    #OneSegNoun.lex(c2n, 'couple')
    #output_filename = "heads_c2n_couple.xlsx"
    
    print(f'Nombre total occurrences de têtes : {OneSegNoun.TOTAL_HEAD_OCC:5d}')
    print(f'Nombre total lemmes de têtes      : {OneSegNoun.TOTAL_HEAD_LEM:5d}')
    print(f'Nombre total de domaine           : {OneSegNoun.TOTAL_DOM:5d}')
    print(OneSegNoun.errors)
    
    if produce_disc_excel:
        OneSegNoun.disciplinary()

    if produce_neo_disc:
        OneSegNoun.disciplinary_neo()
    
    # Disc & Trans
    if produce_quick:
        f = open('output_' + str(OneSegNoun.TOTAL_HEAD_LEM) + '.txt', mode='w', encoding='utf8')
        trans = OneSegNoun.select(med_dom=0.0010)
        # Disc
        for kn, n in OneSegNoun.NOUNS.items():
            for bestd in n.disc_dom:
                if bestd not in disc:
                    disc[bestd] = []
                disc[bestd].append(n)
                if n not in disc_flat:
                    disc_flat.append(n)
        f.write('-----------------------\n Global Info on Disc\n-----------------------\n')
        f.write(f'Number of domains: {len(disc)}\n')
        f.write(f'Number of nb_heads disc: {len(disc_flat)}\n')
        for kd in sorted(SWEET, key=SWEET.get):
            if kd not in disc: continue
            #if kd in ['NONE', '1.shs.autre']: continue
            d = OneSegNoun.DOMAINS[kd]
            total = 0
            total_without_trans = 0
            for n in disc[kd]:
                # Warning:
                # nb_head in noun is for a number of occ of heads! ex: 1736 occ of head "study"
                # nb_head in domain is for the number of lemma of heads! ex: 340 lemma of heads
                total += n.nb_head[kd]
                if n not in trans:
                    total_without_trans += n.nb_head[kd]
            f.write(f"{sweet(kd):48s} {d.filter2:4d}={len(disc[kd]):4d} / {d.nb_heads:5d} {total:5d} or {total_without_trans:5d} / {d.total_heads:5d} {round(total/d.total_heads*100,0)}% {round(total_without_trans/d.total_heads*100,0)}%\n")
        # Trans
        in_tutin = 0
        shared = 0
        f.write('-----------------------\n Info on Disc\n-----------------------\n')
        for kd in sorted(SWEET, key=SWEET.get):
            if kd not in disc: continue
            f.write(sweet(kd) + '\n')
            for n in disc[kd]:
                f.write(f"{n.lemma}\n")
        f.write('-----------------------\n Global Info on Trans\n-----------------------\n')
        f.write("Lemma          Med     Tutin Disc\n")
        for n in trans:
            f.write(f"{n.lemma:15} {n.med_dom:7.6f} {n.lemma in TUTIN} {n in disc_flat}\n")
            if n.lemma in TUTIN: in_tutin += 1
            if n in disc_flat: shared +=1
        f.write("---------- Lemma only ----------\n")
        for n in trans:
            f.write(f"{n.lemma}\n")
        f.write(f'In TUTIN= {in_tutin} / {len(trans)}\n')
        f.write(f'Shared= {shared} / {len(trans)}\n')
        f.close()

    if produce_trans_excel:
        OneSegNoun.to_sheet(output_filename)
    
    if not just_load:
        #Word.write_unknown_lemma()
        filter_titles(debug)
        if debug:
            print('[INFO] Filtered with filter =')
            print(filter_text)
            print('[INFO] --- Total Titles filtered :', len(titles))
        calc_stats(titles)
        if debug: print('[INFO] --- Stats calculated, access by stats')
        if debug:
            print('[INFO] --- Ponctuation which is not segment :')
            Data.from_dict(Title.ponct_no_seg).display()
        print()
        t = titles[list(titles.keys())[0]]
        print(t)
        print(repr(t))
        print()
        print('Root pos :')
        stats["root_pos"].display()
        print('Root lemma :')
        stats['root_lemma'].display(with_line=False, until_total_percent=50.00)
        print('Lemma seg :')
        stats["seg_lemma"].display()
        print('Combi of seg :')
        stats["seg_combi"].display()
        print()
        stats['seg2_nb_dep_from_root'].display()
        stats['seg2_pos_dep_from_root'].display()
        stats['seg2_dep_dep_from_root'].display()
        stats['seg2_lem_dep_from_root'].display(with_line=False, until_total_percent=50.0)
            
if __name__ == '__main__':
    init(True)
