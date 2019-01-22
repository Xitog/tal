"""
XML Glozz Project
------------------------------------------------------------

Evaluating two set of annotation.

Damien Gouteux - 2019 - CC 3.0 BY-SA-NC
"""

#-------------------------------------------------------------------------------
# Imports
#-------------------------------------------------------------------------------

# External library
from lxml import etree
try:
    import openpyxl
except ModuleNotFoundError:
    EXCEL = False
else:
    EXCEL = True
try:
    from nltk.metrics.agreement import AnnotationTask
    from nltk.metrics import ConfusionMatrix
except ModuleNotFoundError:
    KAPPA = False
else:
    KAPPA = True

# Project library
from log import Log

#-------------------------------------------------------------------------------
# Object model
#-------------------------------------------------------------------------------

class Unit:
    """An annotation."""

    PRECISION = 2
    
    def __init__(self, author, date, typ, start, end, features = None):
        self.author = author
        self.date = date
        self.typ = typ
        self.start = start
        self.end = end
        self.features = features if features is not None else {}
    
    def pos_match(self, other):
        if not isinstance(other, Unit):
            raise Exception('Impossible to compare Annotation to ' + str(type(other)))
        return abs(self.start - other.start) <= Unit.PRECISION and \
            abs(self.end - other.end) <= Unit.PRECISION

    def match_on(self, other, feature):
        if not isinstance(other, Unit):
            raise Exception('Impossible to compare Annotation to ' + str(type(other)))
        if feature is None:
            return self.match_all(other)
        else:
            return self.features[feature] == other.features[feature]
    
    def match_all(self, other):
        if not isinstance(other, Unit):
            raise Exception('Impossible to compare Annotation to ' + str(type(other)))
        for key in self.features:
            if self.features[key] != other.features[key]:
                return False
        return True
    
    def __eq__(self, other):
        if not isinstance(other, Unit):
            raise Exception('Impossible to compare Annotation to ' + str(type(other)))
        return self.author == other.author and self.date == other.date and \
               self.typ == other.typ and self.start == other.start and \
               self.end == other.end and self.features == other.features

    def get(self, feature):
        if feature is None:
            s = ""
            nb = 0
            for key,val in self.features.items():
                s += f"{key} : {val}"
                if nb < len(self.features) - 1: s += ','
                nb += 1
            return s
        return self.features[feature] if feature in self.features else None
    
    def __str__(self):
        fun = self.features['fonction'] if 'fonction' in self.features else 'no'
        return f"{self.start}, {self.end}, {fun}"

    def __repr__(self):
        return str(self)


class Annotation:
    """A representation of a Glozz Annotation file (AA)."""

    def __init__(self, filepath : str, code : str):
        self.code = code
        try:
            file = open(filepath, mode='r', encoding='utf8')
        except (FileNotFoundError, IOError) as e:
            Log.error(f'On opening {filepath} : ' + str(e))
        try:
            tree = etree.parse(file)
        except RuntimeError as e:
            Log.error(f'On decoding {filepath} : ' + str(e))
        else:
            file.close()
        root = tree.getroot()
        self.units = []
        for unit in root.iterfind('unit'):
            author = unit.find('metadata/author').text
            date = unit.find('metadata/creation-date').text
            typ = unit.find('characterisation/type').text
            start = int(unit.find('positioning/start/singlePosition').attrib['index'])
            end = int(unit.find('positioning/end/singlePosition').attrib['index'])
            fset = unit.find('characterisation/featureSet')
            features = {}
            for feat in list(fset):
                features[feat.attrib['name']] = feat.text
            self.units.append(Unit(author, date, typ, start, end, features))

    def get_nb_layout(self):
        return sum(1 for u in self.units if u.typ not in ['Signature', 'Mention'])
    
    def get_nb_signature(self):
        return sum(1 for u in self.units if u.typ == 'Signature')

    def get_nb_mention(self):
        return sum(1 for u in self.units if u.typ == 'Mention')

    def get_code(self):
        return self.code
    
    def __len__(self):
        return len(self.units)


class Comparison:
    """Comparison of two annotations."""
    
    def __init__(self, ano1 : Annotation, ano2 : Annotation):
        """Constructor for Comparison instance"""
        self.ano1 = ano1
        self.ano2 = ano2
        self.ano = []
        self.last = None
    
    def diff(self, feature=None):
        """Make a diff between two annotations on a specific feature or all"""
        cpt = 0
        taken = []
        self.ano = []
        self.last = feature
        self.corresponding = 0
        self.matching = 0
        self.no_corresponding = 0
        if self.ano1.get_nb_mention() > self.ano2.get_nb_mention():
            biggest = self.ano1
            smallest = self.ano2
        while cpt < len(biggest.units):
            u1 = biggest.units[cpt]
            if u1.typ != 'Mention':
                cpt += 1
                continue
            found = False
            for u2 in smallest.units:
                if u1.pos_match(u2) and u2 not in taken:
                    self.corresponding += 1
                    match = u1.match_on(u2, feature)
                    self.ano.append([min(u1.start, u2.start), u1, u2, match])
                    if match: self.matching += 1
                    taken.append(u2)
                    found = True
                    break
            if not found:
                self.ano.append([u1.start, u1, None, 'missing annotation'])
            cpt += 1
        # Check if we took everything from the smallest annotation
        for u2 in smallest.units:
            if u2.typ == 'Mention' and u2 not in taken:
                self.ano.append([u2.start, None, u2, 'missing annotation'])
        # Sort
        return self.ano.sort(key=lambda e: e[0])
    

    def to_file(self, ano=None):
        """Produce a simple text file of results"""
        if ano is None:
            ano = self.ano
        datafile = open('15075.ac', mode='r', encoding='utf8')
        data = datafile.read()
        datafile.close()
        last = self.last if self.last is not None else 'all'
        out = open(last + '_diff.txt', mode='w', encoding='utf8')
        out.write("Study on feature : " + last + "\n")
        out.write("===================================================\n\n")
        out.write("Number of units by types\n")
        out.write("------------------------\n")
        out.write(f"Layout types : {self.ano1.get_nb_layout():3d} {self.ano2.get_nb_layout():3d}\n")
        out.write(f"Signature    : {self.ano1.get_nb_signature():3d} {self.ano2.get_nb_signature():3d}\n")
        out.write(f"Mention      : {self.ano1.get_nb_mention():3d} {self.ano2.get_nb_mention():3d}\n")
        out.write(f"-----------------------\n")
        out.write(f"Total        : {len(self.ano1)} {len(self.ano2)}\n")
        out.write("\n")
        out.write("Number of Mentions...\n")
        out.write("--------------------------------------------------\n")
        out.write(f"  with corresponding positions                 {self.corresponding:3d}\n")
        out.write(f"     including X with matching feature         {self.matching:3d}\n")
        out.write(f"     including X with no matching feature      {self.corresponding-self.matching:3d}\n")
        out.write(f"  with no corresponding positions              {self.ano1.get_nb_mention()-self.corresponding + self.ano2.get_nb_mention()-self.corresponding:3d}\n")
        out.write(f"     including X from ano1 {self.ano1.code}                 {self.ano1.get_nb_mention()-self.corresponding:3d}\n")
        out.write(f"     including X from ano2 {self.ano2.code}                 {self.ano2.get_nb_mention()-self.corresponding:3d}\n")
        out.write("\n")
        out.write(f"Mentions with matching feature and corresponding positions    ({self.matching:3d})\n")
        out.write("-------------------------------------------------------------------\n\n")
        nb_common = 1
        for elem in ano:
            u1 = elem[1]
            u2 = elem[2]
            if u1 is not None and u2 is not None and u1.match_on(u2, 'fonction'):
                out.write(f"= {nb_common:03d}\n")
                out.write(f"= {self.ano1.get_code()} {u1.start:6d}, {u1.end:6d}, {data[u1.start : u1.end]:30s}, {u1.typ:10s}, {u1.get(self.last):10s}\n")
                out.write(f"= {self.ano2.get_code()} {u2.start:6d}, {u2.end:6d}, {data[u2.start : u2.end]:30s}, {u2.typ:10s}, {u2.get(self.last):10s}\n\n")
                nb_common += 1
        out.write(f"Mentions with corresponding positions but no matching feature   ({self.corresponding-self.matching:3d})\n")
        out.write("---------------------------------------------------------------------\n\n")
        nb_common = 1
        for elem in ano:
            u1 = elem[1]
            u2 = elem[2]
            if u1 is not None and u2 is not None and not u1.match_on(u2, 'fonction'):
                out.write(f"! {nb_common:03d}\n")
                out.write(f"! {self.ano1.get_code()} {u1.start:6d}, {u1.end:6d}, {data[u1.start : u1.end]:30s}, {u1.typ:10s}, {u1.get(self.last):10s}\n")
                out.write(f"! {self.ano2.get_code()} {u2.start:6d}, {u2.end:6d}, {data[u2.start : u2.end]:30s}, {u2.typ:10s}, {u2.get(self.last):10s}\n\n")
                nb_common += 1
        out.write(f"Mentions with no corresponding positions and no matching feature    ({self.ano1.get_nb_mention()-self.corresponding + self.ano2.get_nb_mention()-self.corresponding:3d})\n")
        out.write("-------------------------------------------------------------------------\n\n")
        nb_uncommon = 1
        for elem in ano:
            u1 = elem[1]
            u2 = elem[2]
            if u1 is not None and u2 is not None:
                continue
            out.write(f"? {nb_uncommon:03d}\n")
            if u1 is None:
                out.write(f"? {self.ano1.get_code()} No mention at this position.\n")
            else:
                out.write(f"? {self.ano1.get_code()} {u1.start:6d}, {u1.end:6d}, {data[u1.start : u1.end]:30s}, {u1.typ:10s}, {u1.get(self.last):10s}\n")
            if u2 is None:
                out.write(f"? {self.ano2.get_code()} No mention at this position.\n\n")
            else:
                out.write(f"? {self.ano2.get_code()} {u2.start:6d}, {u2.end:6d}, {data[u2.start : u2.end]:30s}, {u2.typ:10s}, {u2.get(self.last):10s}\n\n")
            nb_uncommon += 1
        out.write("\nKAPPA\n")
        out.write("-----\n")
        out.write(f"Kappa = {self.kappa()}\n")
        out.close()

    
    def to_excel(self):
        """Produce an Excel file of results"""
        if not EXCEL: return
        datafile = open('15075.ac', mode='r', encoding='utf8')
        data = datafile.read()
        datafile.close()
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.active
        # Units
        if ws is None:
            ws = wb.create_sheet("Units")
        else:
            ws.title = "Units"
        for pair in self.ano:
            u1 = pair[1]
            u2 = pair[2]
            if u1 is not None:
                ws.append([u1.start, u1.end, data[u1.start : u1.end], u1.typ, data[max(0,u1.start-10):min(u1.end+10, len(data)-1)]])
            if u2 is not None:
                ws.append([u2.start, u2.end, data[u2.start : u2.end], u2.typ, data[max(0,u2.start-10):min(u2.end+10, len(data)-1)]])
        # Corresponding units with matching feature
        ws = wb.create_sheet("Matching feature")
        ws.append(['=', 'Nb', 'U1 start', 'U1 end', 'Text', 'Type', self.last if self.last is not None else 'All', \
                             'U2 start', 'U2 end', 'Text', 'Type', self.last if self.last is not None else 'All'])
        nb = 1
        for pair in self.ano:
            u1 = pair[1]
            u2 = pair[2]
            if u1 is not None and u2 is not None and u1.match_on(u2, 'fonction'):
                ws.append([ '=', nb,
                    u1.start, u1.end, data[u1.start : u1.end], u1.typ, u1.get(self.last),
                    u2.start, u2.end, data[u2.start : u2.end], u2.typ, u2.get(self.last)
                    ])
                nb += 1
        # Corresponding units with no matching feature
        ws = wb.create_sheet("No matching feature")
        ws.append(['!', 'Nb', 'U1 start', 'U1 end', 'Text', 'Type', self.last if self.last is not None else 'All', \
                             'U2 start', 'U2 end', 'Text', 'Type', self.last if self.last is not None else 'All'])
        nb = 1
        for pair in self.ano:
            u1 = pair[1]
            u2 = pair[2]
            if u1 is not None and u2 is not None and not u1.match_on(u2, 'fonction'):
                ws.append([ '!', nb,
                    u1.start, u1.end, data[u1.start : u1.end], u1.typ, u1.get(self.last),
                    u2.start, u2.end, data[u2.start : u2.end], u2.typ, u2.get(self.last)
                    ])
                nb += 1
        # Units with no corresponding positions
        ws = wb.create_sheet("No corresponding pos")
        ws.append(['?', 'Nb', 'U1 start', 'U1 end', 'Text', 'Type', self.last if self.last is not None else 'All', \
                             'U2 start', 'U2 end', 'Text', 'Type', self.last if self.last is not None else 'All'])
        nb = 1
        for pair in self.ano:
            u1 = pair[1]
            u2 = pair[2]
            if u1 is None or u2 is None:
                row = ['?', nb]
                if u1 is None:
                    row.extend(['xxx', 'xxx', 'no corresponding unit', 'xxx', 'xxx'])
                else:
                    row.extend([u1.start, u1.end, data[u1.start : u1.end], u1.typ, u1.get(self.last)])
                if u2 is None:
                    row.extend(['xxx', 'xxx', 'no corresponding unit', 'xxx', 'xxx'])
                else:
                    row.extend([u2.start, u2.end, data[u2.start : u2.end], u2.typ, u2.get(self.last)])
                ws.append(row)
                nb += 1
        # Kappa
        ws = wb.create_sheet("Kappa")
        ws.append(['KAPPA', self.kappa()])
        # Save
        last = self.last if self.last is not None else 'all'
        # If the excel is opened we will get an PermissionError.
        # To prevent the loss of our data, the add a modifier to the filename,
        # until finding one not already taken.
        done = False
        modifier = ''
        count = 0
        while not done:
            try:
                # We could add a datetime stamp to the filename
                wb.save(filename=last + '_diff' + modifier + '.xlsx')
            except PermissionError:
                count += 1
                modifier = '_' + str(count)
            else:
                done = True
        return

    
    def info(self):
        """Display information on standard output"""
        Log.info(f'Length of annotation 1: {len(self.ano1.units)}')
        Log.info(f'Length of annotation 2: {len(self.ano2.units)}')
        Log.info(f'Length of fusion      : {len(self.ano)}')


    def kappa(self):
        """Data is a list of list.
           Each element is a list :
           [annotator, element, label]
        """
        if not KAPPA: return 'Not installed'
        #if self.last is None: return # must be specific to a feature
        data = []
        nb = 1
        for elem in self.ano:
            u1 = elem[1]
            u2 = elem[2]
            if u1 is None or u2 is None:
                continue
            else:
                data.append([self.ano1.get_code(), nb, u1.get(self.last)])
                data.append([self.ano2.get_code(), nb, u2.get(self.last)])
                nb += 1
        task = AnnotationTask(data)
        return task.kappa()

#-------------------------------------------------------------------------------
# Main & tool function
#-------------------------------------------------------------------------------

def produce(c : Comparison, feat=None):
    c.diff(feat)
    c.info()
    try:
        c.to_excel()
    except Exception as e:
        print(e)
        Log.warn("Failed to produce Excel output.")
    try:
        c.to_file()
    except Exception as e:
        print(e)
        Log.warn("Failed to produce file output.")

if __name__ == '__main__':
    Log.start()
    Log.info('Production mode')
    dgx = Annotation('15075_dgx.aa', 'DGX')
    slv = Annotation('15075_silvia.aa', 'SLV')
    c = Comparison(dgx, slv)
    produce(c, 'fonction')
    produce(c, 'autoref')
    produce(c)
    Log.end()

