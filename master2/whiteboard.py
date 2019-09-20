# Whiteboard : for dynamic code
# Use reload(wb) to reload this script after having executed datamodel.py
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

#-----------------------------------------------------------
# Final
#-----------------------------------------------------------

def count(dic, key, val=1):
    if key in dic:
        dic[key] += val
    else:
        dic[key] = val


def output_list(wb, title, lst):
    ws = wb.create_sheet(title)
    for val in lst:
        row = []
        for elem in val:
            if type(elem) == tuple:
                for e in elem:
                    row.append(e)
            else:
                row.append(elem)
        ws.append(row)


def output(wb, title, dic):
    ws = wb.create_sheet(title)
    cumul = 0
    tt = sum(dic.values())
    for key in sorted(dic, key=dic.get, reverse=True):
        if type(key) == tuple:
            row = [*key]
        else:
            row = [key]
        row.append(dic[key])
        per = dic[key] / tt
        row.append(per)
        cumul += per
        row.append(cumul)
        ws.append(row)


def is_first_ddaa_nottrans(t): # contraint de la prép à de + contraint de fin en -tion, -ment, -age, -sion
    tt    = None
    p     = None
    nc    = None
    # first pattern
    for i, w in enumerate(t.words):
        if i == 0:
            if w.pos == 'DET' and i + 1 < len(t.words):
                if t.words[i + 1].pos == 'NC' and t.words[i + 1].lemma not in TRANS:
                    tt = t.words[i + 1]
            elif w.pos == 'NC' and w.lemma not in TRANS:
                tt = w
        elif tt is not None and w.pos in ['P', 'P+D']:
            if w.lemma == 'de':
                p = w
            else:
                break # the prep must be de !
        elif p is not None and w.pos == 'NC':
            if w.lemma.endswith('tion') or w.lemma.endswith('ment') or w.lemma.endswith('age') or w.lemma.endswith('sion'):
                nc = w
            else:
                break # the first noun after the prep must be like this
        if tt is not None and p is not None and nc is not None:
            break
    return (tt, p, nc)


def is_first_ddaa(t): # contraint de la prép à de + contraint de fin en -tion, -ment, -age, -sion
    tt    = None
    p     = None
    nc    = None
    # first pattern
    for i, w in enumerate(t.words):
        if i == 0:
            if w.pos == 'DET' and i + 1 < len(t.words):
                if t.words[i + 1].pos == 'NC' and t.words[i + 1].lemma in TRANS:
                    tt = t.words[i + 1]
            elif w.pos == 'NC' and w.lemma in TRANS:
                tt = w
        elif tt is not None and w.pos in ['P', 'P+D']:
            if w.lemma == 'de':
                p = w
            else:
                break # the prep must be de !
        elif p is not None and w.pos == 'NC': # and w.lemma not in TRANS:
            if w.lemma.endswith('tion') or w.lemma.endswith('ment') or w.lemma.endswith('age') or w.lemma.endswith('sion'):
                nc = w
            else:
                break # the first noun after the prep must be like this
        if tt is not None and p is not None and nc is not None:
            break
    return (tt, p, nc)


#-------------------------------------------------------------------------------


# reload(wb) ; cpt, heads = wb.recount_transhead(titles)
def recount_transhead(data):
    nb_app = 0
    errors = {
        '?Approche::NPP'    : 'approche',
        '?Effet::NPP'       : 'effet',
        '?Treatment::NC'    : 'traitement',
        }
    cpt = { 'mono' : 0, 'bi-one-first-seg' : 0, 'bi-one-second-seg' : 0, 'bi-two' : 0 }
    heads = {}
    for kt, t in data.items():
        if t.domain in ['NONE', '1.shs.autre']: continue
        if len(t.roots) == 1:
            r = t.words[t.roots[0]]
            if r.lemma + '::' + r.pos in errors:
                lemma = errors[r.lemma + '::' + r.pos]
                pos == 'NC'
            else:
                lemma = r.lemma
                pos = r.pos
            if lemma in TRANS and pos == 'NC':
                cpt['mono'] += 1
                if lemma not in heads:
                    heads[lemma] = 1
                else:
                    heads[lemma] += 1
                if lemma == 'approche':
                    nb_app += 1
                    if nb_app > 8: continue
        else:
            r1 = t.words[t.roots[0]]
            r2 = t.words[t.roots[1]]
            if r1.lemma + '::' + r1.pos in errors:
                lemma1 = errors[r1.lemma + '::' + r1.pos]
                pos1 = 'NC'
                if lemma1 == 'approche':
                    nb_app += 1
                    if nb_app > 8: continue
            else:
                lemma1 = r1.lemma
                pos1 = r1.pos
            if r2.lemma + '::' + r2.pos in errors:
                lemma2 = errors[r2.lemma + '::' + r2.pos]
                pos2 = 'NC'
                if lemma2 == 'approche':
                    nb_app += 1
                    if nb_app > 8: continue
            else:
                lemma2 = r2.lemma
                pos2 = r2.pos
            if lemma1 in TRANS and pos1 == 'NC' and lemma2 in TRANS and pos2 == 'NC':
                cpt['bi-two'] += 1
                if lemma1 not in heads:
                    heads[lemma1] = 1
                else:
                    heads[lemma1] += 1
                if lemma2 not in heads:
                    heads[lemma2] = 1
                else:
                    heads[lemma2] += 1
            elif lemma1 in TRANS and pos1 == 'NC':
                cpt['bi-one-first-seg'] += 1
                if lemma1 not in heads:
                    heads[lemma1] = 1
                else:
                    heads[lemma1] += 1
            elif lemma2 in TRANS and pos2 == 'NC':
                cpt['bi-one-second-seg'] += 1
                if lemma2 not in heads:
                    heads[lemma2] = 1
                else:
                    heads[lemma2] += 1
    return cpt, heads


# duplicate of datamodel.py
def is_seg(word): # : ; . ? ! + variantes du .?!
     return word.pos == 'PONCT' and word.form in [
         ':', '.', '?', '!', '...', '…', ';', '..', '....', '?.', '?!',
         '...?', '?...', '.....', '!...', '!?', '!.', '!!!', '!!',
         '......', '??', '?..', '.?', '?!...']


# reload(wb) ; r = wb.go(titles)
def go(data):
    nc = {}
    for kt, t in data.items():
        res = is_first(t)
        if res is not None:
            if res['nc'] not in nc:
                nc[res['nc']] = 1
            else:
                nc[res['nc']] += 1
    for k in sorted(nc, key=nc.get, reverse=True):
        print(f"{k:15} {nc[k]:6d}")
    return nc


# reload(wb) ; cpt, heads = wb.count(titles)
def count(data):
    cpt = {'first' : 0, 'second' : 0, 'both' : 0}
    heads = {}
    for kt, t in data.items():
        first = is_first(t)
        second = is_second(t)
        if first is not None and second is not None:
            cpt['both'] += 1
            if first['tt'] not in heads:
                heads[first['tt']] = 1
            else:
                heads[first['tt']] += 1
        elif first is not None:
            cpt['first'] += 1
            if first['tt'] not in heads:
                heads[first['tt']] = 1
            else:
                heads[first['tt']] += 1
        elif is_second(t) is not None:
            cpt['second'] += 1
            if second['tt'] not in heads:
                heads[second['tt']] = 1
            else:
                heads[second['tt']] += 1
    return cpt, heads


# INIT [DET] TRANSHEAD (P à sur de [DET]) | P+D à, de ) NC
def is_first(t):
    i  = t.roots[0]
    r  = t.words[i]
    tt = None
    p  = None
    nc = None
    if i not in [0, 1] or r.lemma not in TRANS: return None
    tt = r
    for j in range(i + 1, len(t.words)):
        nx = t.words[j]
        if p is None:
            if nx.pos == 'ADJ':
                continue
            elif (nx.pos == 'P' and nx.lemma in ['à', 'sur', 'de']) or (nx.pos == 'P+D' and nx.lemma in ['à', 'de']):
                p = nx
                continue
            else:
                return None
        elif nc is None:
            if p.pos == 'P' and nx.pos == 'DET':
                continue
            elif nx.pos == 'NC':
                nc = nx
            else:
                return None
    if p is None or nc is None:
        return None
    else:
        return {'tt' : tt.lemma, 'p' : p.lemma, 'nc' : nc.lemma}


# reload(wb) ; wb.test_is_first()
def test_is_first():
    MW = MockWord
    t = MockTitle([MW('NC', 'problème'), MW('P', 'de'), MW('NC', 'action')], [0])
    print('>>>', is_first(t))
    t = MockTitle([MW('DET', 'le'), MW('NC', 'problème'), MW('P', 'de'), MW('NC', 'action')], [1])
    print('>>>', is_first(t))
    t = MockTitle([MW('DET', 'le'), MW('NC', 'problème'), MW('P', 'de'), MW('DET', 'le'), MW('NC', 'action')], [1])
    print('>>>', is_first(t))
    t = MockTitle([MW('DET', 'le'), MW('NC', 'problème'), MW('ADJ', 'petit'), MW('P', 'de'), MW('NC', 'action')], [1])
    print('>>>', is_first(t))


# SEG [DET] TRANSHEAD (P à sur de [DET]) | P+D à, de ) NC
def is_second(t):
    if len(t.roots) == 1: return None
    i  = t.roots[1]
    r  = t.words[i]
    tt = None
    p  = None
    nc = None
    if i >= 2 and t.words[i - 1].pos == 'DET' and is_seg(t.words[i - 2]):
        pass
    elif i >= 1 and is_seg(t.words[i - 1]):
        pass
    else:
        return None
    if r.lemma not in TRANS: return None
    tt = r
    for j in range(i + 1, len(t.words)):
        nx = t.words[j]
        if p is None:
            if nx.pos == 'ADJ':
                continue
            elif (nx.pos == 'P' and nx.lemma in ['à', 'sur', 'de']) or (nx.pos == 'P+D' and nx.lemma in ['à', 'de']):
                p = nx
                continue
            else:
                return None
        elif nc is None:
            if p.pos == 'P' and nx.pos == 'DET':
                continue
            elif nx.pos == 'NC':
                nc = nx
            else:
                return None
    if p is None or nc is None:
        return None
    else:
        return {'tt' : tt.lemma, 'p' : p.lemma, 'nc' : nc.lemma}


# reload(wb) ; wb.test_is_second()
def test_is_second():
    MW = MockWord
    t = MockTitle([MW('PONCT', ':'), MW('NC', 'problème'), MW('P', 'de'), MW('NC', 'action')], [0, 1])
    print('>>>', is_second(t))
    t = MockTitle([MW('PONCT', ':'), MW('DET', 'le'), MW('NC', 'problème'), MW('P', 'de'), MW('NC', 'action')], [0, 2])
    print('>>>', is_second(t))
    t = MockTitle([MW('PONCT', ':'), MW('DET', 'le'), MW('NC', 'problème'), MW('P', 'de'), MW('DET', 'le'), MW('NC', 'action')], [0, 2])
    print('>>>', is_second(t))
    t = MockTitle([MW('PONCT', ':'), MW('DET', 'le'), MW('NC', 'problème'), MW('ADJ', 'petit'), MW('P', 'de'), MW('NC', 'action')], [0, 2])
    print('>>>', is_second(t))


#-------------------------------------------------------------------------------
    
def is_second_ddaa(t):
    nc1 = MockWord("W", "W") #None # on bazarde
    ponct = None
    tt = None
    p = None
    nc2 = None
    for w in t.words:
        #print(nc1, ponct, tt, p, nc2)
        #if ponct is None and w.pos == 'NC' and w.lemma not in TRANS:
        #    nc1 = w
        #elif nc1 is not None and ponct is None and w.pos == 'PONCT':
        #    ponct = w
        if ponct is None and w.pos == 'PONCT':
            if len(t.segments) > 0 and index(t, w.idw) == t.segments[0]: # NEW: MUST BE THE SEG
                ponct = w
        elif ponct is not None and tt is None and w.pos not in ['DET', 'ADJ', 'NC']:
            break
        elif ponct is not None and tt is None and w.pos == 'NC':
            if w.lemma in TRANS:
                tt = w # must be the first noun after the ponct!
            else:
                break
        elif tt is not None and w.pos in ['P', 'P+D']:
            if w.lemma == 'de':
                p = w
            else:
                break # the prep must be de !
        elif p is not None and w.pos == 'NC': # bazarde no TRANS ici
            if w.lemma.endswith('tion') or w.lemma.endswith('ment') or w.lemma.endswith('age') or w.lemma.endswith('sion'):
                nc2 = w
                break
            else:
                break # the first noun after the prep must be like this
        if nc1 is not None and ponct is not None and tt is not None and p is not None and nc2 is not None:
            break
    return(nc1, ponct, tt, p, nc2)


def is_second_ddaa_nottrans(t):
    nc1 = MockWord("W", "W") #None # on bazarde
    ponct = None
    tt = None
    p = None
    nc2 = None
    for w in t.words:
        #print(nc1, ponct, tt, p, nc2)
        #if ponct is None and w.pos == 'NC' and w.lemma not in TRANS:
        #    nc1 = w
        #elif nc1 is not None and ponct is None and w.pos == 'PONCT':
        #    ponct = w
        if ponct is None and w.pos == 'PONCT':
            if len(t.segments) > 0 and index(t, w.idw) == t.segments[0]: # NEW: MUST BE THE SEG
                ponct = w
        elif ponct is not None and tt is None and w.pos not in ['DET', 'ADJ', 'NC']:
            break
        elif ponct is not None and tt is None and w.pos == 'NC' and w.lemma not in TRANS:
            tt = w
        elif tt is not None and w.pos in ['P', 'P+D']:
            if w.lemma == 'de':
                p = w
            else:
                break # the prep must be de !
        elif p is not None and w.pos == 'NC': # bazarde no TRANS ici
            if w.lemma.endswith('tion') or w.lemma.endswith('ment') or w.lemma.endswith('age') or w.lemma.endswith('sion'):
                nc2 = w
                break
            else:
                break # the first noun after the prep must be like this
        if nc1 is not None and ponct is not None and tt is not None and p is not None and nc2 is not None:
            break
    return(nc1, ponct, tt, p, nc2)


class MockWord:
    def __init__(self, pos, lemma):
        self.pos = pos
        self.lemma = lemma
        self.form = lemma

    def __repr__(self):
        return str(self)
        
    def __str__(self):
        return f"({self.lemma} {self.pos})"
    
class MockTitle:
    def __init__(self, words, iroot):
        self.words = words
        self.roots = iroot


# reload(wb) ; wb.minitest(titles, ['problème'])
def minitest(data, head):
    global TRANS
    old = TRANS
    TRANS = head
    quick_count(data, first=True, notrans=False, onlyhead=True)
    quick_count(data, first=False, notrans=False, onlyhead=True)
    TRANS = old


# reload(wb) ; wb.quick_count(titles, first=, notrans=)
# reload(wb) ; wb.quick_count(titles, first=True, notrans=False, onlyhead=True) # tête trans
def quick_count(data, first=True, notrans=True, onlyhead=False): # ***
    is_head = 0
    is_not_head = 0
    doms = {}
    for kt, t in data.items():
        if notrans:
            if first:
                tt, p, nc = is_first_ddaa_nottrans(t)
            else:
                if len(t.roots) > 1: # only on biseg
                    o, ponct, tt, p, nc = is_second_ddaa_nottrans(t)
                else:
                    o, ponct, tt, p, nc = None, None, None, None, None
        else:
            if first:
                tt, p, nc = is_first_ddaa(t)
            else:
                if len(t.roots) > 1: # only on biseg
                    o, ponct, tt, p, nc = is_second_ddaa(t)
                else:
                    o, ponct, tt, p, nc = None, None, None, None, None
        ok = False
        if first and tt is not None and p is not None and nc is not None:
            ok = True
        elif not first and ponct is not None and tt is not None and p is not None and nc is not None:
            ok = True
        if ok:
            head = False
            if first:
                if t.roots[0] == index(t, tt.idw):
                    is_head += 1
                    head = True
                else:
                    is_not_head += 1
            else:
                if len(t.roots) > 1 and t.roots[1] == index(t, tt.idw):
                    is_head += 1
                    head = True
                else:
                    is_not_head += 1
            if t.domain not in ['1.shs.autre', 'NONE']:
                if not onlyhead or (onlyhead and head):
                    if t.domain not in doms:
                        doms[t.domain] = 1
                    else:
                        doms[t.domain] += 1
    print('NSS Is head:', is_head)
    print('NSS Is not head:', is_not_head)
    sdoms = {}
    for k in sorted(doms, key=doms.get, reverse=True):
        sdoms[k] = doms[k]
    return sdoms


IS_HEAD = 0
IS_NOT_HEAD = 0

import math
# reload(wb) ; wb.correl(300_759, 94_734, 15_698) # coef trans & nss
# reload(wb) ; wb.correl(300_759, 206_025, 20_217) # coef non trans & nss
# wb.correl(1_118_481, 300_759 , 35_915) # coef head & nss
def correl(total, nb_trans, nb_trans_nss): # ***
    serie1 = []
    serie2 = []
    for i in range(0, total):
        if nb_trans_nss > 0 and nb_trans > 0:
            serie1.append(1)
            serie2.append(1)
            nb_trans_nss -= 1
            nb_trans -= 1
        elif nb_trans > 0:
            serie1.append(0)
            serie2.append(1)
            nb_trans -= 1
        else:
            serie1.append(0)
            serie2.append(0)
    moy1 = sum(serie1) / len(serie1)
    moy2 = sum(serie2) / len(serie2)
    som1 = 0
    som2x = 0
    som2y = 0
    for i in range(0, total):
        som1 += (serie1[i] - moy1) * (serie2[i] - moy2)
        som2x += (serie1[i] - moy1) ** 2
        som2y += (serie2[i] - moy2) ** 2
    correl = som1 / math.sqrt(som2x * som2y)
    return correl


# TRANS NSS
def correl_trans_nss_old():
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Test')
    total = 300_759
    nb_trans = 94_734
    nb_trans_nss = 15_698
    #nb_nontrans_nss = 20_217
    for i in range(0, total):
        if nb_trans_nss > 0 and nb_trans > 0:
            ws.append([1, 1])
            nb_trans_nss -= 1
            nb_trans -= 1
        elif nb_trans > 0:
            ws.append([1, 0])
            nb_trans -= 1
        else:
            ws.append([0, 0])
    wb.save('correl_trans_nss.xlsx')


# reload(wb) ; wb.final(titles)
# reload(wb) ; wb.final(titles, 'problème')
# reload(wb) ; wb.final(titles, 'problème', True) # contraint with ddaa
# reload(wb) ; wb.final(titles, 'problème', ddaa=True, save=False)
def final(data, target=None, ddaa=False, save=True):
    global IS_HEAD, IS_NOT_HEAD
    SCHEMA1 = {}
    SCHEMA1_TRANS = {}
    SCHEMA1_NC = {}
    SCHEMA1_P = {}
    SCHEMA1_TEXT = []
    SCHEMA2 = {}
    SCHEMA2_P1 = {}
    SCHEMA2_P2 = {}
    SCHEMA2_NC1 = {}
    SCHEMA2_TRANS = {}
    SCHEMA2_PONCT = {}
    SCHEMA2_P = {}
    SCHEMA2_NC2 = {}
    SCHEMA2_TEXT = []
    cpt = 0
    total = 0
    seuil = 10000
    for kt, t in data.items():
        if not ddaa:
            tt, p, nc = is_first(t)
        else:
            tt, p, nc = is_first_ddaa(t)
        if tt is not None and p is not None and nc is not None:
            if target is None or target == tt.lemma:
                key = (tt.lemma, nc.lemma)
                count(SCHEMA1, key)
                count(SCHEMA1_TRANS, tt.lemma)
                count(SCHEMA1_P, p.lemma)
                count(SCHEMA1_NC, nc.lemma)
                if t.roots[0] == index(t, tt.idw):
                    IS_HEAD += 1
                else:
                    IS_NOT_HEAD += 1
                    #print(kt, t)
                if target is not None:
                    SCHEMA1_TEXT.append([p.lemma, nc.lemma, t.idt, t.text])
        # second pattern
        nc1, ponct, tt, p, nc2 = is_second(t)
        if nc1 is not None and ponct is not None and tt is not None and p is not None and nc2 is not None:
            if target is None or target == tt.lemma:
                key = (nc1.lemma, tt.lemma, nc2.lemma)
                keyP1 = (nc1.lemma, tt.lemma)
                keyP2 = (tt.lemma, nc2.lemma)
                count(SCHEMA2, key)
                count(SCHEMA2_P1, keyP1)
                count(SCHEMA2_P2, keyP2)
                count(SCHEMA2_NC1, nc1.lemma)
                count(SCHEMA2_PONCT, ponct.lemma)
                count(SCHEMA2_TRANS, tt.lemma)
                count(SCHEMA2_P, p.lemma)
                count(SCHEMA2_NC2, nc2.lemma)
                if target is not None:
                    SCHEMA2_TEXT.append([nc1.lemma, ponct.lemma, nc2.lemma, t.idt, t.text])
        cpt += 1
        total += 1
        if cpt == seuil:
            print(f"{total:6d} / {len(data):6d}")
            cpt = 0
    print(f"{total:6d} / {len(data):6d}")
    # write results
    if save:
        wb = openpyxl.Workbook(write_only=True)
        output(wb, "Schema INIT TT NC", SCHEMA1)
        output(wb, "TT", SCHEMA1_TRANS)
        output(wb, "NC", SCHEMA1_NC)
        output(wb, "P", SCHEMA1_P)
        output_list(wb, "TEXT1", SCHEMA1_TEXT)
        output(wb, "Schema NC PONCT TT NC", SCHEMA2)
        output(wb, "NC1 TT", SCHEMA2_P1)
        output(wb, "TT NC2", SCHEMA2_P2)
        output(wb, "NC1", SCHEMA2_NC1)
        output(wb, "TT", SCHEMA2_TRANS)
        output(wb, "PONCT", SCHEMA2_PONCT)
        output(wb, "P", SCHEMA2_P)
        output(wb, "BC2", SCHEMA2_NC2)
        output_list(wb, "TEXT2", SCHEMA2_TEXT)
        if target is None:
            wb.save("trans_schema.xlsx")
        else:
            if not ddaa:
                wb.save("trans_" + target + "_schema.xlsx")
            else:
                wb.save("trans_" + target + "_schema_ddaa.xlsx")


#-----------------------------------------------------------
# Tools
#-----------------------------------------------------------

# Only first level children
def get_children(t, word):
    children = []
    for w in t.words:
        if w.gov == word.idw:
            children.append(w)
    return children


def idw(t, p_idw):
    for w in t.words:
        if w.idw == p_idw:
            return w


def index(t, p_idw):
    for i, w in enumerate(t.words):
        if w.idw == p_idw:
            return i


def is_int(w):
    try:
        int(w.form)
        if ',' in w.form: return False # French flotting number
        return True
    except ValueError:
        return False


patterns = [
    PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid'), # yellow
    PatternFill(start_color='0087CEEB', end_color='0087CEEB', fill_type='solid'), # blue sky
    PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid'), # green
    PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid'), # red
]


# elem pour retrouver les éléments dans le titre
# cs pour trouver les matchs dans tous les titres
# f_cs pour écrire un fichier de résultat

# count nc de nc
# reload(wb) ; wb.count_de(titles, 'all')
def count_de(data, typ='all'):
    TRANS = trans()
    tt = 0
    for kt, t in data.items():
        ok = False
        for w in t.words:
            if w.pos == 'NC':
                children1 = get_children(t, w)
                for w1 in children1:
                    if w1.lemma == 'de':
                        children2 = get_children(t, w1)
                        for w2 in children2:
                            if w2.pos == 'NC':
                                ok = True
                            if ok: break
                    if ok: break
            if ok: break
        if ok and typ == 'all': tt += 1
        elif ok and typ == 'trans' and w.lemma in TRANS: tt += 1
        elif ok and typ == '!trans' and w.lemma not in TRANS: tt += 1
    return tt


# reload(wb) ; wb.count_trans_occ(titles)
def count_trans_occ(data):
    tt_trans = 0
    tt_no_trans = 0
    TRANS = trans()
    for kt, t in data.items():
        for w in t.words:
            if w.pos == 'NC':
                if w.lemma in TRANS:
                    tt_trans += 1
                else:
                    tt_no_trans += 1
    return tt_trans, tt_no_trans


#-----------------------------------------------------------
# CS : NGSS [être] NC
#-----------------------------------------------------------

def elem_n_est_n(t):
    res  = None
    ngss = None
    etre = None
    nco  = None
    for w in t.words:
        if w.pos == 'V' and w.lemma == 'être':
            etre = w
            # il doit y avoir un nom sujet
            children_of_etre = get_children(t, w)
            for w2 in children_of_etre:
                if w2.dep == 'suj' and w2.pos == 'NC':
                    ngss = w2
                elif w2.dep == 'obj' and w2.pos == 'NC':
                    nco = w2
                if ngss is not None and nco is not None:
                    break
        if ngss is not None and nco is not None:
            break
    return (ngss, etre, nco)


# reload(wb) ; res = wb.cs_n_est_n(titles); wb.f_cs_n_est_n(res)
def f_cs_n_est_n(data):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Out')
    all_ngss = {}
    all_nco = {}
    all_cs = {}
    for kt, t in data.items():
        ngss = None
        row = [int(kt)]
        elem = elem_n_est_n(t)
        for w in t.words:
            if w in elem:
                cpt = elem.index(w)
                c = WriteOnlyCell(ws, value=w.form)
                c.fill = patterns[cpt]
                if cpt == 0:
                    ngss = w.lemma
                    if w.lemma not in all_ngss:
                        all_ngss[w.lemma] = 1
                    else:
                        all_ngss[w.lemma] += 1
                elif cpt == 1:
                    pass
                elif cpt == 2:
                    nco = w.lemma
                    if w.lemma not in all_nco:
                        all_nco[w.lemma] = 1
                    else:
                        all_nco[w.lemma] += 1 
                    cs = (ngss, nco)
                    if cs not in all_cs:
                        all_cs[cs] = 1
                    else:
                        all_cs[cs] += 1
                row.append(c)
            else:
                row.append(w.form)
        ws.append(row)
    ws = wb.create_sheet('NGSS')
    ngss_total = sum(all_ngss.values())
    for k in sorted(all_ngss, key=all_ngss.get, reverse=True):
        ws.append([k, all_ngss[k], all_ngss[k] / ngss_total])
    ws = wb.create_sheet('NCO')
    nco_total = sum(all_nco.values())
    for k in sorted(all_nco, key=all_nco.get, reverse=True):
        ws.append([k, all_nco[k], all_nco[k] / nco_total])
    ws = wb.create_sheet('CS')
    cs_total = sum(all_cs.values())
    for k in sorted(all_cs, key=all_cs.get, reverse=True):
        ws.append([*k, all_cs[k], all_cs[k] / cs_total])
    ws = wb.create_sheet('Info')
    ws.append(['Total occ NGSS', ngss_total])
    ws.append(['Total occ NCO', nco_total])
    ws.append(['Total diff CS', cs_total])
    ws.append(['Total res', len(data)])
    wb.save('CS-n_est_n.xlsx')


def cs_n_est_n(data):
    res = {}
    for kt, t in data.items():
        ok_suj = False
        ok_obj = False
        for w in t.words:
            if w.pos == 'V' and w.lemma == 'être':
                # il doit y avoir un nom sujet
                children_of_etre = get_children(t, w)
                for w2 in children_of_etre:
                    if w2.dep == 'suj' and w2.pos == 'NC':
                        ok_suj = True
                    elif w2.dep == 'obj' and w2.pos == 'NC':
                        ok_obj = True
                    if ok_suj and ok_obj:
                        break
            if ok_suj and ok_obj:
                res[kt] = t
                break
    return res


#-----------------------------------------------------------
# CS : NGSS [être] que
#-----------------------------------------------------------


def elem_que(t):
    res  = None
    ngss = None
    etre = None
    que  = None
    for w in t.words:
        # on a un que CS
        if w.pos == 'CS' and w.lemma == 'que':
            que = w
            i_que = index(t, w.idw)
            # il y a avant un verbe être conjugué à l'indicatif
            if i_que > 1 and t.words[i_que - 1].lemma == 'être' and t.words[i_que - 1].pos == 'V':
                etre = t.words[i_que - 1]
                # il doit y avoir un nom sujet
                children_of_etre = get_children(t, t.words[i_que - 1])
                for w2 in children_of_etre:
                    if w2.dep == 'suj' and w2.pos == 'NC':
                        ngss = w2
                        res = (ngss, que)
                        break
                    elif w2.form == "c'": # le problème, c'est que ...
                        i_c = index(t, w2.idw)
                        if i_c > 2 and t.words[i_c - 1].form == ',' and t.words[i_c - 2].pos == 'NC':
                            ngss = t.words[i_c - 2]
                            res = (ngss, que)
                            break
                        elif i_c > 1 and t.words[i_c - 1].pos == 'NC':
                            ngss = t.words[i_c - 1]
                            res = (ngss, que)
                            break
            # il y a un nom avant
            elif i_que > 1 and t.words[i_que - 1].pos == 'NC':
                ngss = t.words[i_que - 1]
                res = (ngss, que, etre)
                break
    return res


# reload(wb) ; res = wb.cs_que(titles); wb.f_cs_que(res)
def f_cs_que(data):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Out')
    all_ngss = {}
    for kt, t in data.items():
        ngss = None
        row = [int(kt)]
        elem = elem_que(t)
        for w in t.words:
            if w in elem:
                cpt = elem.index(w)
                c = WriteOnlyCell(ws, value=w.form)
                c.fill = patterns[cpt]
                if cpt == 0:
                    ngss = w.lemma
                    if w.lemma not in all_ngss:
                        all_ngss[w.lemma] = 1
                    else:
                        all_ngss[w.lemma] += 1
                row.append(c)
            else:
                row.append(w.form)
        ws.append(row)
    ws = wb.create_sheet('NGSS')
    ngss_total = sum(all_ngss.values())
    for k in sorted(all_ngss, key=all_ngss.get, reverse=True):
        ws.append([k, all_ngss[k], all_ngss[k] / ngss_total])
    ws = wb.create_sheet('Info')
    ws.append(['Total occ NGSS', ngss_total])
    ws.append(['Total res', len(data)])
    wb.save('CS-que.xlsx')


def cs_que(data):
    res = {}
    for kt, t in data.items():
        ok = False
        for w in t.words:
            # on a un que CS
            if w.pos == 'CS' and w.lemma == 'que':
                i_que = index(t, w.idw)
                # il y a avant un verbe être conjugué à l'indicatif
                if i_que > 1 and t.words[i_que - 1].lemma == 'être' and t.words[i_que - 1].pos == 'V':
                    # il doit y avoir un nom sujet
                    children_of_etre = get_children(t, t.words[i_que - 1])
                    for w2 in children_of_etre:
                        if w2.dep == 'suj' and w2.pos == 'NC':
                            ok = True
                        elif w2.form == "c'": # le problème, c'est que ...
                            i_c = index(t, w2.idw)
                            if i_c > 2 and t.words[i_c - 1].form == ',' and t.words[i_c - 2].pos == 'NC':
                                ngss = t.words[i_c - 2]
                                ok = True
                                break
                            elif i_c > 1 and t.words[i_c - 1].pos == 'NC':
                                ngss = t.words[i_c - 1]
                                ok = True
                                break
                # il y a un nom avant
                elif i_que > 1 and t.words[i_que - 1].pos == 'NC':
                    ok = True
            if ok:
                res[kt] = t
                break
    return res


#-----------------------------------------------------------
# CS : NGSS [être] de inf
#-----------------------------------------------------------
# cs_de_inf(data)   fait un filtrage sur les titres qui correspondent à ce pattern
# elem_cs_de_inf(t) récupère les mots clés de la CS dans un titre
# f_cs_de_inf(data) met les résultats filtrés dans un fichier Excel à l'aide de elem_

def elem_cs_de_inf(t):
    res  = None
    ngss = None
    etre = None
    de   = None
    inf  = None
    for w in t.words:
        if w.pos == 'VINF':
            inf = w
            parent1 = idw(t, w.gov)
            if parent1 is not None and parent1.lemma == 'de' and parent1.pos == 'P':
                de = parent1
                parent2 = idw(t, parent1.gov)
                if parent2 is not None and parent2.lemma == 'être' and parent2.pos == 'V':
                    etre = parent2
                    children2 = get_children(t, parent2)
                    for w2 in children2:
                        if w2.dep == 'suj' and w2.pos == 'NC':
                            ngss = w2
                            res = (ngss, de, inf, etre)
                            break
                elif parent2 is not None and parent2.pos == 'NC':
                    ngss = parent2
                    res = (ngss, de, inf)
                    break
    if res is not None:
        # voyons s'il y a un NC juste avant le de pour corriger les dépendances
        i_de = index(t, de.idw)
        if i_de > 0 and t.words[i_de - 1].pos == 'NC':
            if len(res) == 3:
                res = (t.words[i_de - 1], de, inf)
            elif len(res) == 4:
                print(t.idt)
                res = (t.words[i_de - 1], de, inf, etre)
    return res


# reload(wb) ; res = wb.cs_de_inf(titles); wb.f_cs_de_inf(res)
def f_cs_de_inf(data):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Out')
    all_ngss = {}
    all_inf = {}
    all_cs = {}
    for kt, t in data.items():
        ngss = None
        row = [int(kt)]
        elem = elem_cs_de_inf(t)
        for w in t.words:
            if w in elem:
                cpt = elem.index(w)
                c = WriteOnlyCell(ws, value=w.form)
                c.fill = patterns[cpt]
                if cpt == 0:
                    ngss = w.lemma
                    if w.lemma not in all_ngss:
                        all_ngss[w.lemma] = 1
                    else:
                        all_ngss[w.lemma] += 1
                elif cpt == 2: # on ne retient pas le "être" optionnel
                    cs = (ngss, 'de', w.lemma)
                    if w.lemma not in all_inf:
                        all_inf[w.lemma] = 1
                    else:
                        all_inf[w.lemma] += 1
                    if cs not in all_cs:
                        all_cs[cs] = 1
                    else:
                        all_cs[cs] += 1
                row.append(c)
            else:
                row.append(w.form)
        ws.append(row)
    ws = wb.create_sheet('NGSS')
    ngss_total = sum(all_ngss.values())
    for k in sorted(all_ngss, key=all_ngss.get, reverse=True):
        ws.append([k, all_ngss[k], all_ngss[k] / ngss_total])
    ws = wb.create_sheet('INF')
    inf_total = sum(all_inf.values())
    for k in sorted(all_inf, key=all_inf.get, reverse=True):
        ws.append([k, all_inf[k], all_inf[k] / inf_total])
    ws = wb.create_sheet('CS')
    cs_total = sum(all_cs.values())
    for k in sorted(all_cs, key=all_cs.get, reverse=True):
        ws.append([*k, all_cs[k], all_cs[k] / cs_total])
    ws = wb.create_sheet('Info')
    ws.append(['Total occ NGSS', ngss_total])
    ws.append(['Total occ CS', cs_total])
    ws.append(['Total res', len(data)])
    wb.save('CS-inf.xlsx')


# ne pas est un children du verbe
# le sujet est un children du verbe
def cs_de_inf(data):
    res = {}
    for kt, t in data.items():
        ok = False
        for w in t.words:
            # on a un infinitf
            if w.pos == 'VINF' and w.form not in ['Grégoire', 'Alexandre'] and not is_int(w):
                # suppression de bien-être
                if w.lemma == 'être':
                    i_etre = index(t, w.idw)
                    if i_etre > 1 and t.words[i_etre - 1].form == '-' and t.words[i_etre - 2].form == 'bien':
                        continue
                # a-t-il pour parent de ?
                parent1 = idw(t, w.gov)
                if parent1 is not None and parent1.lemma == 'de' and parent1.pos == 'P':
                    # en vue de est un no go
                    i_de = index(t, parent1.idw)
                    if i_de > 0 and t.words[i_de - 1].form == 'vue':
                        break
                    # a-t-il pour parent un être ?
                    parent2 = idw(t, parent1.gov)
                    if parent2 is not None and parent2.lemma == 'être' and parent2.pos == 'V':
                        # si oui, son sujet doit être un NC !
                        children2 = get_children(t, parent2)
                        for w2 in children2:
                            if w2.dep == 'suj' and w2.pos == 'NC':
                                ok = True
                    elif parent2 is not None and parent2.pos == 'NC':
                        # si non, son recteur doit être le NGSS
                        ok = True
            if ok:
                res[kt] = t
                break
    return res

#-----------------------------------------------------------
# DISC & TRANS
#-----------------------------------------------------------


def read_disc_from_output_xlsx(filename, debug=True):
    TRANS = trans()
    filename = filename + '.xlsx' if not filename.endswith('.xlsx') else filename
    wb = load_workbook(filename, read_only=True)
    disc_heads = {}
    for i, sn in enumerate(wb.sheetnames):
        if i >= 4: # disc
            disc = sn[sn.index('.') + 2:]
            disc_heads[disc] = []
            ws = wb[sn]
            for row in ws.rows:
                key = str(row[1].value) + '::' + str(row[2].value)
                disc_heads[disc].append(key)
    disc_filtered = {}
    all_lemma_filtered = []
    for disc in disc_heads:
        disc_filtered[disc] = []
        for key in disc_heads[disc]:
            lemma, pos = key.split('::') # place existe en NC et NPP pour '1.shs.infocom'
            if lemma in TRANS:
                print(f"{disc:15} {lemma}")
                break
            disc_filtered[disc].append(lemma)
            if lemma not in all_lemma_filtered: all_lemma_filtered.append(lemma)
    if debug:
        print('Discipline      NbHead Filter - Minus Kept %')
        for disc in disc_heads:
            ld = len(disc_heads[disc])
            lf = len(disc_filtered[disc])
            print(f"{disc:15} {ld:6d} {lf:6d} -{ld-lf:6d} {lf/ld*100:5.2f}%")
    return disc_heads, all_lemma_filtered
rdfo = read_disc_from_output_xlsx


def recouvrement(DISC=None, TRANS=None):
    if type(DISC) != list: DISC = disc(DISC)
    if type(TRANS) != list: TRANS = trans(TRANS)
    recouv = 0
    for t in TRANS:
        if t in DISC:
            recouv += 1
    print(f"Length of DISC = {len(DISC)} Length of TRANS = {len(TRANS)}")
    print(f"Recouv : {recouv} / {len(TRANS)} {recouv/len(TRANS)*100:5.2f}")


def disc():
    file = open(r'.\output\disc_heads.txt', mode='r', encoding='utf8')
    lines = file.readlines()
    file.close()
    res = []
    for lin in lines:
        res.append(lin.rstrip())
    return res


def trans():
    file = open(r'.\output\trans_heads.txt', mode='r', encoding='utf8')
    lines = file.readlines()
    file.close()
    res = []
    for lin in lines:
        res.append(lin.rstrip())
    return res
get_trans = trans
TRANS = get_trans()


def select_trans_t(data):
    res = {}
    TRANS = trans()
    for kt, t in data.items():
        root1 = t.words[t.roots[0]]
        if root1.lemma in TRANS:
            res[kt] = t
            continue
        elif len(t.roots) > 1:
            root2 = t.words[t.roots[1]]
            if root2.lemma in TRANS:
                res[kt] = t
    return res


#-----------------------------------------------------------

# wb.search(titles, 'problème', text='Le problème')
def search(data, root1, root2=None, nb=10, text=None):
    for kt, t in data.items():
        if root2 is not None and len(t.roots) < 2: continue
        if text is not None and text not in t.text: continue
        troot1 = t.words[t.roots[0]]
        if root2 is None:
            if root1 == troot1.lemma:
                print(t.idt, t)
                nb -= 1
        else:
            troot2 = t.words[t.roots[1]]
            if root1 == troot1.lemma and root2 == troot2.lemma:
                print(t.idt, t)
                nb -= 1
        if nb == 0: break

# no title with same root2 and 'problème' AND 'question' as root1
def xsearch(data):
    for kt, t in data.items():
        if len(t.roots) < 2: continue
        troot2 = t.words[t.roots[1]]
        tp = None
        tq = None
        for kkt, tt in data.items():
            troot1 = t.words[t.roots[0]]
            if troot1.lemma == 'problème':
                tp = tt
            if troot1.lemma == 'question':
                tq = tt
        if tp is not None and tq is not None:
            print(tp)
            print(tq)


# wb.on_each(t121, wb.find_sub_roots)
def on_each(titles, f):
    for kt, t in titles.items():
        f(t)


# show in which segment are the roots
# already in the attribute roots_by_segment
def root_in_seg(titles, with_subroots = False):
    res = {}
    for kt, t in titles.items():
        if len(t.segments) > 1:
            if len(t.segments) == 2:
                # if the second segmentator mark is a closing one it's ok
                if t.segments[-1] != t.len_with_ponct - 1:
                    raise Exception('Too many segmentors: ' + str(t.segments))
            else:
                raise Exception('Too many segments: ' + str(len(t.segments)))
        # everything is ok
        seg_mark = t.segments[0]
        if len(t.roots) == 0:
            raise Exception('No root')
        elif len(t.roots) > 2:
            raise Exception('Too many roots: ' + str(len(t.roots)))
        in_seg1 = 0
        in_seg2 = 0
        if with_subroots == True:
            roots = t.roots + t.subroots
        else:
            roots = t.roots
        for root in roots:
            if root < seg_mark:
                in_seg1 += 1
            elif root > seg_mark:
                in_seg2 += 1
            else:
                raise Exception('Root is segmentator!')
        key = (in_seg1, in_seg2)
        if key not in res:
            res[key] = 1
        else:
            res[key] += 1
    return res


def pprint(data, nb_max=None):
    total = 0
    for k, v in data.items():
        total += v
    nb = 0
    for key in sorted(data, key=data.get, reverse=True):
        val = data[key]
        print(f"{str(key):20} {val:10} {round(val/total * 100, 2):5.2f}")
        nb += 1
        if nb_max is not None and nb == nb_max:
            break


def deter_for(titles, domains, lem1, lem2, seg):
    xlem = {
        'le' : 'def',
        'la' : 'def',
        'les' : 'def',
        'un' : 'indef',
        'une' : 'indef',
    }
    dets = { 'none' : 0}
    for kt, t in titles.items():
        if t.domain not in domains: continue
        root1 = t.words[t.roots[0]]
        root2 = t.words[t.roots[1]]
        if root1.lemma != lem1 or root2.lemma != lem2: continue
        found = False
        if seg == 1:
            start = 0
            end = t.segments[0]
            r = root1
            ir = t.roots[0]
        elif seg == 2:
            start = t.segments[0] + 1
            end = len(t.words)
            r = root2
            ir = t.roots[1]
        else:
            raise Exception('Number of seg not handled')
        for iw in range(start, end):
            w = t.words[iw]
            #print(t.roots[seg], w.pos, w.gov, w.dep)
            if w.pos == 'DET' and w.gov == r.idw and w.dep == 'det':
                if iw < ir:
                    lem = w.lemma
                    if lem in xlem:
                        lem = xlem[lem]
                    if lem not in dets:
                        dets[lem] = 1
                    else:
                        dets[lem] += 1
                    found = True
                    break
        if not found:
            dets['none'] += 1
    return dets


def go_deter(data, dom): # OneSegNoun.DOMAINS
    print('rôle', deter_for(data, dom, 'rôle', 'cas', 1))
    print('cas', deter_for(data, dom, 'rôle', 'cas', 2))
    print('approche', deter_for(data, dom, 'approche', 'cas', 1))
    print('cas', deter_for(data, dom, 'approche', 'cas', 2))
    print('apport', deter_for(data, dom, 'apport', 'exemple', 1))
    print('exemple', deter_for(data, dom, 'apport', 'exemple', 2))
    print('effet', deter_for(data, dom, 'effet', 'cas', 1))
    print('cas', deter_for(data, dom, 'effet', 'cas', 2))


# deprecated
def deter(titles, domains):
    det = {}
    no_det = 0
    nb_found = {}
    for kt, t in titles.items():
        if t.domain not in domains: continue
        root = t.roots[0] + 1
        found = 0
        for iw, w in enumerate(t.words):
            if w.pos == 'DET' and w.gov == root and w.dep == 'det':
                if iw < root:
                    found +=1
                    if w.form in det:
                        det[w.form] += 1
                    else:
                        det[w.form] = 1
                else:
                    pass # Dîtes-le avec des cartes
                    #t.info()
                    #raise Exception("Deter AFTER root " + str(w))
        if found in nb_found:
            nb_found[found] += 1
        else:
            nb_found[found] = 1
            #print('-------------------')
            #print('Root :', root)
            #t.info()
    ordered_det = {}
    for key in sorted(det, key=det.get, reverse=True):
        val = det[key]
        ordered_det[key] = val
    return ordered_det, nb_found


def comple(titles):
    dep = {}
    no_dep = 0
    nb_found = {}
    for kt, t in titles.items():
        root = t.roots[0] + 1
        found = 0
        for iw, w in enumerate(t.words):
            if w.pos == 'P' and w.gov == root and w.dep == 'dep':
                if iw >= root:
                    found +=1
                    if w.form.lower() in dep:
                        dep[w.form.lower()] += 1
                    else:
                        dep[w.form.lower()] = 1
                else:
                    t.info()
                    raise Exception("Compl BEFORE root " + str(w))
        if found in nb_found:
            nb_found[found] += 1
        else:
            nb_found[found] = 1
            #print('-------------------')
            #print('Root :', root)
            #t.info()
    ordered_dep = {}
    for key in sorted(dep, key=dep.get, reverse=True):
        val = dep[key]
        ordered_dep[key] = val
    return ordered_dep, nb_found


def nb_seg_2_nb_restarts(titles):
    count = 0
    exceptions = []
    for kt, t in titles.items():
        if t.nb_seg >= (t.nb_restarts + 1):
            count += 1
        else:
            exceptions.append(t)
    return count, count / len(titles), len(exceptions), exceptions



# reload(wb);t=wb.before_restart(titles)
def before_restart(titles):
    all_before_roots = {} 
    for kt, t in titles.items():
        if t.nb_restarts != 1:
            continue
        i = t.restarts[0]
        before = t.words[i - 1]
        key = before.form + ' : ' + before.pos + ' : ' + before.lemma
        count(all_before_roots, key)
    for k in sorted(all_before_roots, key=all_before_roots.get, reverse=True):
        print(f"{k:22} {all_before_roots[k]:10d}")
    return all_before_roots


def test(titles):
    maxt = 10
    cpt = 0
    tx = []
    for kt, t in titles.items():
        for w in t.words:
            if w.gov == 0 and w.dep in ['_', 'root'] and w.pos == 'NC':
                print(t.idt, t.text)
                cpt += 1
                tx.append(t)
        if cpt == maxt:
            break
    return tx


def test1(titles):
    governed = {}
    "On regarde si la root est un P ou P+D, on regarde qui dépend d'elle"
    for kt, t in titles.items():
        for w in t.words:
            if w.gov == 0 and w.dep in ['_', 'root'] and w.pos in ['P', 'P+D']:
                nbgov = 0
                for ww in t.words:
                    if w.gov == 1:
                        nbgov += 1
                if nbgov in governed:
                    governed[nbgov] += 1
                else:
                    governed[nbgov] = 1
    return governed
