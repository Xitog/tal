#===============================================================================
# Code handling for handling data in Excel format XLSX
#-------------------------------------------------------------------------------
# Author : Damien Gouteux
# Last updated : 08 April 2018
# Technologies : Python, Excel XSLX
# Usage :
#     Write or read data in an Excel XLSX file.
#===============================================================================

#-------------------------------------------------------------------------------
# Import
#-------------------------------------------------------------------------------

# external
import openpyxl
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
#import xlwt

#-------------------------------------------------------------------------------
# Classes
#-------------------------------------------------------------------------------

class MiniCell:

    yellow = PatternFill(fgColor='FFFFFF00', fill_type='solid')
    
    def __init__(self, val, bg=None, w=None):
        self.val = val
        self.bg = bg
        self.w = w

    def to_cell(self, ws):
        if self.bg is None and self.w is None:
            return self.val
        else:
            cell = WriteOnlyCell(ws, self.val)
            if self.bg is not None:
                cell.fill = self.bg
            return cell


class ExcelFile:

    def __init__(self, name, mode='r'):
        self.name = name
        self.mode = mode
        if mode == 'w':
            self.wb = openpyxl.Workbook(write_only=True)
        elif mode == 'r':
            self.wb = openpyxl.load_workbook(name, read_only=True) # keep_vba=True

    # find a sheet from its name or its number
    # ex : sheet('TITLE') or sheet(0)
    def sheet(self, ws):
        if isinstance(ws, str):
            if ws not in self.wb.sheetnames:
                raise Exception("Sheet unknown: ", ws)
            return self.wb[ws]
        elif isinstance(ws, int):
            if len(self.wb.sheetnames) <= ws:
                raise Exception("Not enough sheet: ", ws, " (", len(self.wb.sheetnames), ")")
            ws = self.wb.sheetnames[ws]
            return self.wb[ws]
    
    def save_to_sheet(self, name, values, order_col=0, reverse_order=True, percent_col=None):
        ws = self.wb.create_sheet(name)
        if name == 'TITLES': # hack
            column = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
            for c in column:
                ws.column_dimensions[c].width = 2
        if isinstance(values, dict):
            values = values.values()
        if percent_col is not None:
            count = 0
            for list_of_values in values:
                count += list_of_values[percent_col]
        if order_col is not None:
            for list_of_values in sorted(values, key=lambda t: t[order_col], reverse=reverse_order):
                out = []
                for obj in list_of_values:
                    if isinstance(obj, MiniCell):
                        out.append(obj.to_cell(ws))
                    else:
                        out.append(obj)
                if percent_col is not None:
                    out.append(list_of_values[percent_col] / count)
                ws.append(out)
        else:
            for list_of_values in values:
                out = []
                for obj in list_of_values:
                    if isinstance(obj, MiniCell):
                        out.append(obj.to_cell(ws))
                    else:
                        out.append(obj)
                if percent_col is not None:
                    out.append(list_of_values[percent_col] / count)
                ws.append(out)
    
    def save_to_sheet_old(self, name, values, percent=None, test_val=None):
        # ws = wb.add_sheet(name)
        if self.nb_sheet > 0:
            ws = self.wb.create_sheet(name)
            self.nb_sheet += 1
        else:
            ws= self.wb.active
            ws.title = name
            self.nb_sheet += 1
        row = 1
        for val in sorted(values, key=values.get, reverse=True): #sorted(values.keys()):
            if test_val is None or test_val(values[val]):
                nb = 1
                if type(val) in [tuple, list]:
                    for v in val:
                        ws.cell(column=nb, row=row, value=v)
                        nb += 1
                else:
                    ws.cell(column=nb, row=row, value=val)
                    nb += 1
                #ws.write(row, 0, val)
                if type(values[val]) in [tuple, list]:
                    for v in values[val]:
                        ws.cell(column=nb, row=row, value=v)
                        #ws.write(row, nb, v)
                        nb += 1
                else:
                    ws.cell(column=nb, row=row, value=values[val])
                    #ws.write(row, nb, values[val])
                    nb += 1
                if percent is not None: # percent on last value
                    if type(values[val]) in [tuple, list]:
                        ws.cell(column=nb, row=row, value=values[val][-1] / percent)
                    else:
                        ws.cell(column=nb, row=row, value=values[val] / percent)
                    #ws.write(row, nb, values[val] / percent)
                row += 1

    def load_sheet(self, idx=None, key=0, ignore=None):
        # load any sheet in dict with the choosen column as a key
        # ignoring specific column (a list of column : 0, 1, 2...)
        # by default, loading the sheet number 0, with key=first column, ignoring none
        # test
        if self.mode != 'r':
            raise Exception('This ExcelFile should be on read mode.')
        # loading the sheet
        if idx is None:
            idx = 0
        sheet = self.sheet(idx)
        # fetching the data
        data = {}
        nb_row = 0
        for row in sheet.rows: # sheet.iter_rows(min_row=0):
            nb_cell = 0
            last_id = 0
            values = []
            for cell in row:
                if cell.value is not None:
                    if nb_cell == key:
                        last_id = cell.value
                    else:
                        if ignore is None or nb_cell not in ignore:
                            values.append(cell.value)
                nb_cell += 1
            data[last_id] = values
            nb_row += 1
        return data
    
    def close(self):
        if self.mode != 'w':
            raise Exception('This ExcelFile should be on write mode.')
        done = False
        modifier = ''
        count = 0
        while not done:
            try:
                self.wb.save(filename=self.name + modifier + '.xlsx')
                done = True
            except PermissionError:
                count += 1
                modifier = '_' + str(count)


class DynMatrix:
    
    def __init__(self, name):
        self.name = name
        self.words = []
        self.count = []
        self.oriented_cooccurrences = {}
        self.matrix = None

    def info(self):
        print('Words:', len(self.words))
        count = 0
        for w1, vals in self.oriented_cooccurrences.items():
            print('   ', f'{w1:10}', vals)
            for w2, nb in vals.items():
                count += nb
        print('Oriented Cooccurrences:', count, '\n')
        
    def add(self, w1, w2):
        # count each words
        for w in [w1, w2]:
            if w not in self.words:
                self.words.append(w)
                self.count.append(1)
            else:
                i = self.words.index(w)
                self.count[i] += 1
        # save the oriented cooccurrence
        if w1 not in self.oriented_cooccurrences:
            self.oriented_cooccurrences[w1] = {w2 : 1}
        else:
            if w2 in self.oriented_cooccurrences[w1]:
                self.oriented_cooccurrences[w1][w2] += 1
            else:
                self.oriented_cooccurrences[w1][w2] = 1
    
    def filter(self, threshold, display=False):
        new_words = []
        new_count = []
        for i in range(len(self.words)):
            if self.count[i] >= threshold:
                new_words.append(self.words[i])
                new_count.append(0)
            elif display:
                print('Deleting', self.words[i], 'occ=', self.count[i], '\n')
        for word in self.words:
            if word not in new_words:
                if word in self.oriented_cooccurrences:
                    del self.oriented_cooccurrences[word]
                for other, oc in self.oriented_cooccurrences.items():
                    if word in oc:
                        del oc[word]
        print('Filtered:', len(new_words), 'on', len(self.words), 'with threshold=', threshold)
        self.words = new_words
        self.count = new_count
        for w1, others in self.oriented_cooccurrences.items():
            i = self.words.index(w1)
            for w2, nb in others.items():
                j = self.words.index(w2)
                self.count[i] += nb
                self.count[j] += nb
        self.build_matrix(decorated=True)
    
    def build_matrix(self, decorated=False):
        # build structure
        if decorated:
            content = [ ['MATRIX'] + self.words + ['Total'] ]
        else:
            content = []
        for w1 in range(len(self.words)):
            if decorated:
                content.append([self.words[w1]])
            else:
                content.append([])
            for w2 in range(len(self.words)):
                content[-1].append(0)
        # fill structure
        for w1, oc in self.oriented_cooccurrences.items():
            for w2 in oc:
                i = self.words.index(w1)
                j = self.words.index(w2)
                if decorated:
                    i += 1
                    j += 1
                content[i][j] += oc[w2]
        # count
        for irow, row in enumerate(content):
            count = 0
            for icell, cell in enumerate(row):
                if (irow > 0 and decorated) and icell > 0:
                    count += cell
            if irow > 0:
                row.append(count)
        # filter
        new_content = []
        for index, row in enumerate(content):
            if index == 0 and decorated:
                    new_content.append(row)
            elif row[-1] > 0:
                new_content.append(row)
        self.matrix = new_content
        return self.matrix
    
    def to_excel(self, regen=False, decorated=False, debug=False, excel=None):
        if regen or self.matrix is None: self.build_matrix(decorated)
        if excel is None:
            excel = ExcelFile(name = self.name, mode = 'w')
        excel.save_to_sheet(name = 'matrix', values = self.matrix, order_col=None)
        rows = []
        for i in range(len(self.words)):
            rows.append([self.words[i], self.count[i]])
        excel.save_to_sheet(name = 'count', values = rows, order_col=1, reverse_order=True)
        excel.close()
        del excel
        if debug:
            f = open(self.name + '_debug', mode='w', encoding='utf-8')
            for w1, oc in self.oriented_cooccurrences.items():
                for w2 in oc:
                    f.write(w1 + ',' + w2 + ',' + str(oc[w2]))
            f.close()
    
    def display(self, regen=False, decorated=False):
        if regen or self.matrix is None: self.build_matrix(decorated)
        nb_row = 0
        for row in self.matrix:
            for cell in row:
                print(f"{str(cell):10}", end='')
            print()
            nb_row += 1
        print()

#-------------------------------------------------------------------------------
# Test code if main
#-------------------------------------------------------------------------------

def test_excel():
    excel = ExcelFile(name='test', mode='w')
    data = {
        "bordeaux" : ['Bordeaux', 'préfecture', 33, 'Gironde'],
        "toulouse" : ['Toulouse', 'préfecture', 31, 'Haute-Garonne'],
        "albi"     : ['Albi', 'préfecture', 81, 'Tarn-et-Garonne'],
        "castres"  : ['Castres', 'sous-préfecture', 81, 'Tarn-et-Garonne'],
    }
    excel.save_to_sheet('VILLES', data, order_col = 2, reverse_order = False)
    excel.close()

def test_dynmatrix():
    dm = DynMatrix('Test')
    dm.add('blanche', 'noire')
    dm.add('noire', 'jaune')
    dm.add('jaune', 'jaune')
    dm.add('bleu', 'rouge')
    dm.add('rouge', 'bleu')
    dm.add('blanche', 'bleu')
    dm.add('jaune', 'bleu')
    dm.add('bleu', 'violet')
    dm.add('jaune', 'bleu')
    dm.add('jaune', 'bleu')
    dm.info()
    dm.display(decorated=True)
    dm.filter(2, display=True)
    dm.display(decorated=True)
    dm.to_excel(decorated=True)

if __name__ == '__main__':
    #test_excel()
    test_dynmatrix()

    
    
